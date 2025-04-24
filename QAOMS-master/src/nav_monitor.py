import shutil

from broker import BrokerApp
from globals.config import OMS_CONFIG
from datetime import datetime, timedelta
from globals.addresses import ADDR
from globals.log_setup import LOG
from markets import US_MARKET
import numpy as np
import openpyxl
from openpyxl.styles import numbers, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import pandas as pd
import shutil
from threading import Thread
import time as sleeper


class NAVMonitor:
    # This object monitors all connected accounts and writes the account nav to a local file at a regular period.

    def __init__(self, oms_instance):

        self.oms = oms_instance
        self.keep_alive = True
        self.current_navs = {}  # key=account ids,
        # value=[Current NAV, % Since Yesterday, % this month, % this year, last updated]

        self.file_path = os.path.join(ADDR.folder_performance, 'NAV and Performance.xlsx')

        self._last_half_hour = US_MARKET.current_time - timedelta(seconds=3600)

        thread = Thread(target=self.monitor_navs, name='NAV Monitor')
        thread.start()

    def monitor_navs(self):

        while self.keep_alive:

            accounts_to_write = []
            now = US_MARKET.current_time
            scheduled_write = now.replace(minute=(0 if now.minute < 30 else 30), second=0, microsecond=0)
            # 2 minute interval - for use in testing:
            # minute = now.minute - now.minute % 2
            # scheduled_write = now.replace(minute=minute, second=0, microsecond=0)

            for a in self.oms.trading_accounts:
                if (a.ib_type == 'FUND ADVISORY' and OMS_CONFIG and OMS_CONFIG.config and
                        not OMS_CONFIG.config['Report on Advisor Account']):
                    # don't write NAV for FUND ADVISORY accounts
                    continue
                elif a.last_nav_write is None or a.last_nav_write < scheduled_write:
                    accounts_to_write.append(a)

            for account in accounts_to_write:

                nav = None
                try:
                    broker = [b for b in BrokerApp._instances if account.id in b.api.connected_accounts]
                    # broker, trade_register = self.oms.retrieve_broker(account)
                    if broker:
                        broker = broker[0]
                        nav = broker.nlv(account_id=account.id,
                                         currency='BASE',
                                         refresh=True)

                        # LOG.warning(f'Initial NAV write complete for {account.alias}')
                except ConnectionError:
                    LOG.warning(f'Unable to retrieve NAV from account {account.alias} ({account.id}) - '
                                f'Connection error')
                except ValueError as e:
                    LOG.warning(f'Unable to retrieve NAV from account {account.alias} ({account.id}) - '
                                f'Value error - {e.args[0]}')
                except Exception as e:
                    LOG.warning(f'Unable to retrieve NAV from account {account.alias} ({account.id}) - '
                                f'Error - {e.args[0]}')
                if nav:
                    if self.export_nav_to_file(account, nav) is True:
                        LOG.debug(f'Account {account.id} NAV recorded as $AUD {nav}.')
                        account.last_nav_write = US_MARKET.current_time
                else:
                    sleeper.sleep(1)

            # sleep interval
            for _ in range(50):
                if self.keep_alive:
                    sleeper.sleep(0.1)
                else:
                    break

    def export_nav_to_file(self, account, nav):
        # this function replaces the deprecated write_nav_to_file

        try:

            # Load the workbook if exists, else create a new one

            if not os.path.isdir(ADDR.folder_performance):
                # cannot access directory; if proceed beyond this point, chance that existing file will be
                # overwritten
                LOG.warning(f'Could not find NAV folder at {ADDR.folder_performance}. NAV export cancelled.')
                return

            wb_exists = os.path.isfile(self.file_path)
            if wb_exists:
                LOG.debug(f'NAV - loading NAV workbook from {self.file_path}.')
                wb = openpyxl.load_workbook(self.file_path)
            else:
                LOG.debug('NAV - creating NAV workbook.')
                wb = openpyxl.Workbook()

            # 1 - LOAD EXISTING DATA TO MEMORY

            sheet_name = account.alias + ' NAV'
            column_names = ['Date', 'NAV', 'Funds In/Out', 'Last Updated', 'Daily Return', 'EoM Return', 'Year Return']

            if sheet_name not in wb.sheetnames:
                LOG.debug(f'NAV - Creating sheet {sheet_name}')
                wb.create_sheet(sheet_name)
                df = pd.DataFrame(columns=column_names)
            else:
                df = pd.read_excel(self.file_path, sheet_name=sheet_name)

            df.rename(columns={'Funds': 'Funds In/Out'}, inplace=True)

            if list(df.columns)[:4] != column_names[:4]:
                LOG.warning(f'Unexpected column names found in {account.alias} NAV sheet: {list(df.columns)[:4]}. NAV '
                            f'export cancelled.')
                return

            df = df[['Date', 'NAV', 'Funds In/Out', 'Last Updated']]
            df['Date'] = pd.to_datetime(df['Date']).dt.date
            df.set_index(df.columns[0], inplace=True)  # set date column as index

            # 2 - ADD CURRENT NAV

            nav = float(nav)
            today = US_MARKET.current_time.date()
            # today_str = datetime.strftime(today, date_format)

            if today in df.index:
                backup = False
                df.at[today, 'NAV'] = nav
            else:
                backup = True  # save daily backups
                df.loc[today] = [nav] + [''] * (len(df.columns) - 1)

            # 3 - ADD FUND CHANGES

            df['Funds In/Out'] = None

            fund_changes = [fc for fc in OMS_CONFIG.config['Funds'] if fc[0] == account.id]
            fund_dates = [fc[1] for fc in fund_changes]

            for fd in fund_dates:
                change_amount = sum([fc[2] for fc in fund_changes if fc[1] == fd])
                if fd in df.index:
                    df.at[fd, 'Funds In/Out'] = change_amount
                else:
                    df.loc[fd] = ['', change_amount, '']

                if df.at[fd, 'NAV'] in ['', None] or np.isnan(df.at[fd, 'NAV']):
                    prev_date = df.index[df.index < fd].max()
                    if pd.notna(prev_date):
                        df.at[fd, 'NAV'] = df.at[prev_date, 'NAV']
                        df.at[fd, 'Last Updated'] = f'Carried forward from {prev_date}'
                    else:
                        earliest_nav = df[pd.notna(df['NAV'])].index.min()
                        LOG.warning(f'{account.alias} - CONFIG FILE shows a deposit/withdrawal on a date ({fd}) prior '
                                    f'to or on the first date that a NAV was recorded for the account ({earliest_nav}).'
                                    f' This will prevent calculation of returns for the associated periods.')

            # 4 - TIMESTAMP
            df.at[today, 'Last Updated'] = US_MARKET.current_time.replace(tzinfo=None)

            # 5 - RECALCULATE
            df.sort_index(ascending=False, inplace=True)
            # Daily return - Time Weighted Return calculation
            f = pd.to_numeric(df['Funds In/Out'], errors='coerce').fillna(0)
            n = pd.to_numeric(df['NAV'], errors='coerce').fillna(0)
            valid = n.notna() & n.shift(-1).notna() & f.notna()
            df['Daily Return'] = np.where(valid, ((n - f) - n.shift(-1)) / n.shift(-1), np.nan)

            # Monthly return, as product of daily returns
            df['Month'] = pd.to_datetime(df.index).to_period('M')
            monthly_return = df.groupby('Month')['Daily Return'].apply(lambda x: (x + 1).prod()) - 1

            df['EoM Return'] = df['Month'].map(monthly_return)  # type: ignore
            first_in_month = ~df['Month'].duplicated()
            df['EoM Return'] = df['EoM Return'].where(first_in_month)

            # Yearly return, as a product of daily returns
            df['Year'] = pd.to_datetime(df.index).year
            yearly_return = df.groupby('Year')['Daily Return'].apply(lambda x: (x + 1).prod()) - 1
            df['Year Return'] = df['Year'].map(yearly_return)
            first_in_year = ~df['Year'].duplicated()
            df['Year Return'] = df['Year Return'].where(first_in_year)

            df.drop(columns=['Month', 'Year'], inplace=True)

            # 6 - EXPORT TO EXCEL, WRITE TO ACCOUNT SHEET, WRITE TO PERFORMANCE SUMMARY SHEET, and FORMAT
            try:
                day_return = df.at[today, "Daily Return"]
                month_return = df.at[today, "EoM Return"]
                year_return = df.at[today, 'Year Return']
            except (ValueError, TypeError):
                day_return = 'error'
                month_return = 'error'
                year_return = 'error'

            # archive existing file if the nav being written is for a new day.
            if backup:
                archive_folder = os.path.join(ADDR.folder_performance, 'ARCHIVE')
                archive_date = US_MARKET.current_time.date() - timedelta(days=1)
                archive_file_path = os.path.join(archive_folder, f'NAV and Performance at {archive_date}.xlsx')
                os.makedirs(archive_folder, exist_ok=True)
                shutil.copy(self.file_path, archive_file_path)

                # keep all archives for end of months, and for the last 30 days
                for f in os.listdir(archive_folder):
                    try:
                        date = datetime.strptime(f.split('NAV and Performance at ')[1].strip().split('.')[0],
                                                 '%Y-%m-%d').date()
                    except (IndexError, ValueError):
                        continue
                    if (date + timedelta(days=1)).day != 1 and date < datetime.now().date() - timedelta(days=30):
                        os.remove(os.path.join(archive_folder, f))

            # --- 6a: Write Account Detail Sheet ---
            if sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])
            ws = wb.create_sheet(sheet_name)

            df_fmt = df.copy()
            df_fmt = df_fmt.replace([np.inf, -np.inf], 'error').fillna('')
            rows = list(dataframe_to_rows(df_fmt, index=True, header=True))
            rows[0][0] = 'Date'
            del rows[1]
            for r in rows:
                ws.append(r)

            for col in ('B', 'C'):
                for cell in ws[col][1:]:
                    cell.number_format = '[$AUD] #,##0;-[$AUD] #,##0'

            for col in ('A', 'B', 'C', 'E', 'F', 'G', 'H'):
                ws.column_dimensions[col].width = 15
            ws.column_dimensions['D'].width = 20

            for col in ('E', 'F', 'G'):
                for cell in ws[col][1:]:
                    cell.number_format = numbers.FORMAT_PERCENTAGE_00

            for col in ['A', 'D']:
                for cell in ws[col][1:]:
                    cell.alignment = Alignment(horizontal='center')

            for row in ws.iter_rows():
                for cell in row:
                    cell.border = Border()

            # --- 6b: Write to PERFORMANCE Sheet ---
            sheet_name = 'PERFORMANCE'
            if sheet_name not in wb.sheetnames:
                LOG.debug('NAV - creating summary performance sheet')
                ws = wb.create_sheet(sheet_name)
                ws.append(['Account', 'NAV', 'Since Yesterday', 'This Month', 'This Year'])
            else:
                ws = wb[sheet_name]

            row_to_update = None
            for row in range(2, ws.max_row + 1):
                if ws.cell(row, 1).value == account.alias:
                    row_to_update = row
                    break
            if not row_to_update:
                row_to_update = 2
                ws.insert_rows(2)
                LOG.debug(f'NAV - adding account {account.alias} to summary in row {row_to_update}')
                ws.cell(row_to_update, 1).value = account.alias

            ws.cell(row_to_update, 2).value = nav
            ws.cell(row_to_update, 2).number_format = '[$AUD] #,##0.00;-[$AUD] #,##0.00'

            for i, result in enumerate([day_return, month_return, year_return]):
                cell = ws.cell(row_to_update, i + 3)
                cell.value = result
                cell.number_format = numbers.FORMAT_PERCENTAGE_00

            wb.save(self.file_path)

            # LOG to screen
            LOG.info(f'{account.alias} ({account.id}) - Current NAV ${nav:.2f}')
            LOG.info(f'{account.alias} ({account.id}) - Day return calculated as {day_return*100:.2f}%')
            LOG.info(f'{account.alias} ({account.id}) - MTD return calculated as {month_return*100:.2f}%')
            LOG.info(f'{account.alias} ({account.id}) - YTD return calculated as {year_return*100:.2f}%')

            return True

        except (FileNotFoundError, PermissionError):
            LOG.warning(f'Could not save NAV to file at {self.file_path}.  Check that the file is not in use/open.')
            return False
        except Exception as e:
            LOG.warning(f'Could not save NAV to file at {self.file_path}. Error: {e.args[0]}')
            return False

    @DeprecationWarning
    def write_nav_to_file(self, account, nav):
        # replaced by export_nav_to_file above.

        date_format = '%Y-%m-%d'

        try:

            # Load the workbook if exists, else create a new one

            if not os.path.isdir(ADDR.folder_performance):
                # cannot access directory; if proceed beyond this point, chance that existing file will be
                # overwritten
                return

            wb_exists = os.path.isfile(self.file_path)
            if wb_exists:
                LOG.debug(f'NAV - loading NAV workbook from {self.file_path}.')
                wb = openpyxl.load_workbook(self.file_path)
            else:
                LOG.debug('NAV - creating NAV workbook.')
                wb = openpyxl.Workbook()

            # 1 - WRITE NAV

            nav = float(nav)
            today = US_MARKET.current_time.date()
            today_str = datetime.strftime(today, date_format)

            sheet_name = account.alias + ' NAV'

            if sheet_name not in wb.sheetnames:
                LOG.debug(f'NAV - Creating sheet {sheet_name}')
                wb.create_sheet(sheet_name)
                ws = wb[sheet_name]
                ws.append(["Date", "NAV", "Funds In/Out", "Last Updated"])  # Add headers to new sheet
            else:
                ws = wb[sheet_name]
                if ws.cell(1, 3).value == 'Last Updated':
                    LOG.debug(f'NAV - Updating legacy format for {sheet_name}')
                    # update legacy format
                    for row in range(2, ws.max_row + 1):
                        ws.cell(row, 3).value = None
                    ws.cell(1, 3).value = 'Funds In/Out'
                    ws.cell(1, 4).value = 'Last Updated'
                for row in range(2, ws.max_row + 1):  # handle previous bug that wrote rows with no date
                    if ws.cell(row, 1).value == '':
                        LOG.debug(f'NAV - Deleting row {row} on {sheet_name}')
                        ws.delete_rows(row)

            # Check if today's date is already in the sheet
            row_to_update = None
            for row in range(2, ws.max_row + 1):
                if ws.cell(row, 1).value == today_str:
                    LOG.debug(f'NAV - found date {today_str} in row {row_to_update} in sheet {sheet_name}')
                    row_to_update = row
                    break

            if not row_to_update:
                # insert a new row for this date; dates are recorded descending
                LOG.debug(f'NAV - inserting row for new date in sheet {sheet_name}')
                ws.insert_rows(2)
                row_to_update = 2

            # write data
            LOG.debug(f'NAV - writing date {today_str} to cell ({row_to_update}, 1) on sheet {sheet_name}')
            ws.cell(row_to_update, 1).value = today_str
            LOG.debug(f'NAV - writing NAV ${nav} to cell ({row_to_update}, 2) on sheet {sheet_name}')
            cell = ws.cell(row_to_update, 2)
            cell.value = nav
            cell.number_format = '[$AUD] #,##0.00;-[$AUD] #,##0.00'
            last_updated = datetime.strftime(US_MARKET.current_time, '%Y-%m-%d %H:%M:%S')
            LOG.debug(f'NAV - writing last updated {last_updated} to cell ({row_to_update}, 4) on sheet {sheet_name}')
            cell = ws.cell(row_to_update, 4)
            cell.value = last_updated

            # 2 - WRITE DEPOSITS & WITHDRAWALS

            fund_changes = [fc for fc in OMS_CONFIG.config['Funds'] if fc[0] == account.id]

            for row in range(2, ws.max_row + 1):
                row_date = ws.cell(row, 1).value
                if row_date:
                    row_date = datetime.strptime(row_date, date_format).date()
                    fc = [fc[2] for fc in fund_changes if fc[1] == row_date]
                    if fc:
                        cell = ws.cell(row, 3)
                        cell.value = fc[0]
                        cell.number_format = '[$AUD] #,##0.00;-[$AUD] #,##0.00'

            # 3 - CALCULATE PERFORMANCE

            # calculate data
            nav_today = None
            nav_yesterday = None
            nav_end_of_prev_month = None
            nav_end_of_prev_year = None
            nav_earliest = None
            yesterday = None
            end_of_prev_month = None
            end_of_prev_year = None
            date_earliest = None

            for row in range(2, ws.max_row + 1):
                try:
                    row_date = datetime.strptime(ws.cell(row, 1).value, date_format).date()
                except (ValueError, TypeError):
                    continue

                row_nav = ws.cell(row, 2).value
                row_fc = sum([fc[2] for fc in fund_changes if fc[1] <= row_date])
                row_nav = row_nav - row_fc

                if date_earliest is None or row_date < date_earliest:  # type: ignore
                    date_earliest = row_date
                    nav_earliest = row_nav

                if row_date == today:
                    nav_today = row_nav
                    continue

                if not yesterday or row_date > yesterday and row_date != today:  # type: ignore
                    yesterday = row_date
                    nav_yesterday = row_nav

                if row_date.year == today.year:
                    if row_date.month == today.month - 1:
                        if not end_of_prev_month or row_date > end_of_prev_month:  # type: ignore
                            end_of_prev_month = row_date
                            nav_end_of_prev_month = row_nav
                elif row_date.year == today.year - 1:
                    if row_date.month == 12 and today.month == 1:
                        if not end_of_prev_month or row_date > end_of_prev_month:  # type: ignore
                            end_of_prev_month = row_date
                            nav_end_of_prev_month = row_nav
                    if not end_of_prev_year or row_date > end_of_prev_year:
                        end_of_prev_year = row_date
                        nav_end_of_prev_year = row_nav

            if nav_today is not None:
                day_return = ((nav_today - nav_yesterday) / nav_yesterday) if nav_yesterday else 0

                month_return = (((nav_today - nav_end_of_prev_month) / nav_end_of_prev_month)
                                if nav_end_of_prev_month else 0)
                if not nav_end_of_prev_month:
                    month_return = ((nav_today - nav_earliest) / nav_earliest) if nav_earliest else 0

                year_return = ((nav_today - nav_end_of_prev_year) / nav_end_of_prev_year) if nav_end_of_prev_year else 0
                if not nav_end_of_prev_year:
                    year_return = ((nav_today - nav_earliest) / nav_earliest) if nav_earliest else 0
            else:
                day_return, month_return, year_return = 0, 0, 0

            LOG.info(f'{account.id} - Current NAV ${nav_today}')
            LOG.info(f'{account.id} - Day return calculated as {round(day_return * 100, 1)}%')
            LOG.info(f'{account.id} - MTD return calculated as {round(month_return * 100, 1)}%')
            LOG.info(f'{account.id} - YTD return calculated as {round(year_return * 100, 1)}%')

            # 4 - WRITE PERFORMANCE

            sheet_name = 'PERFORMANCE'

            if sheet_name not in wb.sheetnames:
                LOG.debug('NAV - creating summary performance sheet')
                wb.create_sheet(sheet_name)
                ws = wb[sheet_name]
                ws.append(["Account", "NAV", "Since Yesterday",  "This Month", "This Year"])  # Add headers to sheet
            else:
                ws = wb[sheet_name]

            # Check if account is in sheet
            row_to_update = None
            for row in range(2, ws.max_row + 1):
                if ws.cell(row, 1).value == account.alias:
                    LOG.debug(f'NAV - found existing account {account.alias} in performance summary sheet')
                    row_to_update = row
                    break
            if not row_to_update:
                ws.insert_rows(2)
                row_to_update = 2
                LOG.debug(f'NAV - adding account {account.alias} to performance summary sheet in row {row_to_update}')
                ws.cell(row_to_update, 1).value = account.alias

            # write data
            cell = ws.cell(row_to_update, 2)
            cell.value = nav
            cell.number_format = '[$AUD] #,##0.00;-[$AUD] #,##0.00'

            for idx, result in enumerate([day_return, month_return, year_return]):
                cell = ws.cell(row_to_update, idx + 3)
                cell.value = result
                cell.number_format = numbers.FORMAT_PERCENTAGE_00

            self.current_navs[account.alias] = [nav, day_return, month_return, year_return, last_updated]

            # Save File

            # double check file does not already exist
            wb.save(self.file_path)

        except (FileNotFoundError, PermissionError):
            LOG.warning(f'Could not save NAV to file at {self.file_path}.  Check that the file is not in use/open.')
        except Exception as e:
            LOG.warning(f'Could not save NAV to file at {self.file_path}. Error: {e.args[0]}')


if __name__ == '__main__':
    from trading_objects.account import Account
    ac = Account(id=0, alias='MIT-ALPHA', ib_type='', ib_algo='', ib_platform='', ib_port=0, ib_client_id=0,
                 strategy_allocations={}, strategy_security_types={}, ignored_positions=[])

    nm = NAVMonitor(None)
    OMS_CONFIG.config['Funds'] = [[0, datetime(2025, 1, 8).date(), 50000],
                           [0, datetime(2024, 7, 23).date(), 50000]]
    nm.export_nav_to_file(ac, 150000)
    os.startfile(nm.file_path)