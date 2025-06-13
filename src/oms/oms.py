__version__ = '0.1.17.5'

'''
oms.py is the main OMS management code.  It maintains the order file monitoring, initiates/maintains broker objects,
places daily exit/entry orders, generates daily report emails and daily RealTest output files.

An 'instance' of this object is created in the gui creation script started at the bottom of this module.  This is 
the entry point for the whole program.

The code has grown as requirement details have developed, and is now too long/ungainly; refactoring is required to 
make the code more readable.
'''

import sys
from pathlib import Path
# Add the project root directory to the Python path
sys.path.append(str(Path(__file__).resolve().parent.parent.parent))

from src.utils import utilities as util  # import first to initiate log
from src.broker.broker import BrokerApp
from copy import deepcopy
import csv
from datetime import datetime, timedelta
from decimal import Decimal
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.base import MIMEBase
from email.header import Header
from email.utils import formataddr
from email import encoders
from src.gui.gui import OMSgui
import logging
from src.config.globals.addresses import ADDR
from src.config.globals.config import OMS_SETTINGS, OMS_CONFIG
from src.config.globals.email_manager import EMAIL_MANAGER
from src.config.globals.log_setup import LOG, customSMTPHandler
from src.config.globals.signals import SIGNALS
from src.config.globals.trading import TGL
from src.markets.markets import US_MARKET
from src.account.nav_monitor import NAVMonitor
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from src.trades.orders.order_file_handler import FileWatcher
from src.trades.orders.orders import OrderRef
import src.trades.orders as o
import os
from os.path import basename
import pandas as pd
from PyQt6.QtWidgets import QApplication
from PyQt6.QtGui import QPixmap
import re
import shutil
from smtplib import SMTP, SMTPDataError, SMTPServerDisconnected, SMTPAuthenticationError
import sys
import threading
import time as sleeper
import traceback
from src.trades.trades import TradeRecord, TradeRegister


class OMS:

    def __init__(self):

        # when the OMS instance is created, these variables are set up
        self.brokers = []            # Broker objects with a connection to the API, one broker object per account
        self.raw_orders = []         # List of valid orders supplied to the OMS by the user; invalid orders rejected
        self.data_broker = None      # the first broker connection is used for accessing market data, contracts, etc.

        # connections
        SIGNALS.raise_oms_error.connect(self.raise_oms_error)

        # settings
        OMS_SETTINGS.load_serialised_settings()
        self.manual_order_process_button_clicked = False

        # and these system/internal variables
        self._reporting_lock = threading.Lock()
        self._broker_lock = threading.Lock()
        self._blocking_function = ''  # used by the gui to determine which process currently holds the reporting lock
        self.ib_stock_margin_data = pd.DataFrame()
        self.ib_stock_borrow_costs = pd.DataFrame()

        self._nav_monitor = NAVMonitor(self)

        print('\n\n')
        self._stop_signal = False  # for use by the OMS to stop operations after an error
        self.keep_alive = True
        self.initialised = False
        self.order_processing_status = ''
        self.version = __version__

    @property
    def config(self):
        return OMS_CONFIG.config

    @property
    def trading_accounts(self):
        return OMS_CONFIG.trading_accounts

    @property
    def strategies(self):
        return OMS_CONFIG.strategies

    @property
    def ignored_positions(self):
        return OMS_CONFIG.ignored_positions

    def raise_oms_error(self, error_tuple):
        level, exception, stop_signal = error_tuple
        # level: int, exception: Exception, stop_signal: bool

        # this handles EXCEPTIONS (only) which result in the OMS or the orders being stopped.
        # Other errors (e.g. an invalid order in the imported orders) are logged directly.
        trace = traceback.format_exception(exception)
        error_text = traceback.format_exception_only(exception)
        SIGNALS.exception.emit((level, error_text, stop_signal, trace))
        self._stop_signal = stop_signal
        if self._stop_signal:
            LOG.warning('OMS Stop Signal set to True')

    def stop_signal(self):
        # waits until the stop_signal flag is cleared, or the oms is forced to exit
        while self._stop_signal and self.keep_alive:
            sleeper.sleep(1)
        if not self.keep_alive:
            sys.exit()
        else:
            return False

    def run(self):

        self.stop_signal()

        # load CONFIG data from config excel file
        loaded = False
        while not loaded and not self.stop_signal():
            LOG.info('Loading config')
            try:
                OMS_CONFIG.load_config()
                loaded = True
            except (ValueError, NotADirectoryError) as error:
                self.raise_oms_error((2, error, True))
            except Exception as error:
                self.raise_oms_error((1, error, True))

        try:
            # This is the main controller that initiates and maintains the OMS's work.  The function:
            #  - connects to the first broker to retrieve market data
            #  - establishes threads which ensure reports are issued daily as required
            #  - continuously watches for a new order file; whenever an order file is found, it is processed and orders
            #  - are placed
            # See notes below for details and more info

            # connect to broker accounts
            LOG.info('Connecting to accounts via Interactive Brokers API.')
            for a in self.trading_accounts:
                self.stop_signal()
                broker, trade_register = self.retrieve_broker(a)
                if broker:
                    LOG.info(f'IB API connection established for {a.ib_type} account {a.alias} ({a.id}).  '
                             f'Trading Mode = {broker.api.trading_mode}.')

            # use the first broker connection to retrieve market data etc.
            while not self.data_broker and not self.stop_signal():
                if self.trading_accounts:
                    LOG.info(f'Initialising market data connection using account '
                             f'{self.trading_accounts[0].alias}')
                    self.data_broker = self.retrieve_broker(self.trading_accounts[0], refresh_data=False)[0]
                else:
                    self.raise_oms_error((2,
                                          ValueError('No valid account available for market data connection'),
                                          True))
            self.stop_signal()
            LOG.info('Retrieving market session data...')
            try:
                US_MARKET.update_market_sessions(self.data_broker)
            except ConnectionError as e:
                self.raise_oms_error((2, e, True))
            LOG.info('Retrieving market session data complete.')

            # initiate thread to watch for Order Files
            self.stop_signal()
            file_watch = FileWatcher()
            LOG.info(f'\r\n\nWaiting for order file. To initiate trading please drag or save an order file to '
                     f'{ADDR.folder_order_files}\n')

            # initiate timers for Daily Reports and RealTest output files
            self.initialised = True
            SIGNALS.start_timers.emit()

            # Infinite loop to monitor for order files and place orders
            while True:
                while not self.stop_signal():

                    OMS_CONFIG.load_config()

                    # wait for the order processing time, or for the user to click the manual orders button
                    if self.is_time_to_process_orders() or self.manual_order_process_button_clicked:

                        self.manual_order_process_button_clicked = False

                        self.order_processing_status = 'Checking for and loading order files'
                        file_watch.check_for_order_files()

                        # reset variables before processing order files
                        for account in self.trading_accounts:
                            account.exit_results = []
                            account.entry_results = []
                            account.runtime_errors = []
                        orders_found = False

                        # retrieve current IB data for short stocks
                        self.order_processing_status = 'Retrieving IB short margin and fee rate data'
                        try:
                            self.ib_stock_margin_data = pd.read_csv(
                                "ftp://shortstock:%20@ftp2.interactivebrokers.com/stockmargin_final_dtls."
                                "IBLLC-US.dat",
                                delimiter='|', skiprows=1)
                        except Exception:
                            self.ib_stock_margin_data = pd.DataFrame()

                        try:
                            self.ib_stock_borrow_costs = pd.read_csv(
                                "ftp://shortstock:%20@ftp2.interactivebrokers.com/usa.txt", delimiter='|',
                                skiprows=1)
                        except Exception as e:
                            self.ib_stock_borrow_costs = pd.DataFrame()

                        # reload account, strategy and system config data before each order import - allows the
                        # user to change settings/add trading systems etc. between placing order files if necessary
                        self.order_processing_status = 'Loading config file'
                        OMS_CONFIG.load_config()

                        ### PROCESS ORDER FILES  ###
                        broker_data_prepped = False
                        while not self.stop_signal() and file_watch.order_data:
                            # the user or another system has saved a new order file in the order files folder

                            orders_found = True

                            if not broker_data_prepped:
                                # do this once (here) per batch of order files, and only when order files provided

                                # reconcile trade register with the broker account and report discrepancies
                                for account in self.trading_accounts:
                                    self.order_processing_status = (f'Reconciling API and trade register for '
                                                                    f'{account.alias}')
                                    broker, trade_register = self.retrieve_broker(account, refresh_data=True)
                                    self.reconcile(broker, account)

                                broker_data_prepped = True

                            # load orders
                            file_name, raw_orders = file_watch.order_data.pop(0)
                            LOG.info(f'{len(raw_orders)} {self.plo(raw_orders)} found in file.\n')

                            # get the new orders from the order file.
                            # See order_file_handler.validate_and_load_order_file() to see how the RT or Amibroker
                            # file is converted to rawOrder objects
                            loop_orders = raw_orders.copy()
                            for order in [ordr for ordr in loop_orders if not ordr.account_id]:
                                # AmiBroker files do not have an account ID associated with each order - each order
                                # should be applied to all accounts
                                mult_account_raw_orders = []
                                for account in self.trading_accounts:
                                    account_order = deepcopy(order)
                                    account_order.account_id = account.id
                                    mult_account_raw_orders.append(account_order)
                                raw_orders.remove(order)
                                raw_orders.extend(mult_account_raw_orders)
                            for order in raw_orders:
                                order.status = 'PENDING'
                            remaining_orders = raw_orders

                            self.raw_orders.extend(raw_orders)
                            SIGNALS.raw_order_update.emit()

                            # for each account:
                            # (1) review raw EXIT orders and place as appropriate
                            # (2) review raw ENTRY orders and place as appropriate
                            # (3) issue log to user by email
                            for account in self.trading_accounts:

                                self.order_processing_status = f'Processing orders for {account.alias}'

                                # note that the imported raw_orders are filtered into account subgroups (here).  An
                                # order file can therefore contain any combination of account orders (one account or
                                # many), and the OMS will only place one account's set of orders at a time.
                                account_orders = [order for order in raw_orders if order.account_id == account.id]
                                self.stop_signal()

                                # are there any orders for this account?  if yes then proceed
                                if account_orders:

                                    broker, trade_register = self.retrieve_broker(account, refresh_data=False)

                                    broker.api.verify_connection(ignore_market_data=True)

                                    account.log.info("#" * shutil.get_terminal_size()[0])
                                    account.log.info(f'Processing raw orders for {account.alias}')
                                    account.log.info(f'Source File:  {file_name}')
                                    account.log.info(f'Order Count:  {len(raw_orders)}')
                                    account.log.info(f'Strategy/ies: {", ".join(list(set([ro.strategy_model_string for ro in raw_orders])))}')
                                    account.log.info('-' * 20)
                                    account.log.info(f'{account.alias} NLV = {broker.base_currency(account.id)} '
                                                     f'{broker.nlv(account.id, refresh=True)} '
                                                     f'(BASE)')
                                    account.log.info(f'{account.alias} NLV = USD {broker.nlv(account.id, "USD")}')
                                    account.log.info("#" * shutil.get_terminal_size()[0])

                                    # have we found a valid broker and trade_register for this account? if yes, proceed
                                    if broker and trade_register and not self.stop_signal():

                                        # record here which orders are left over after we place this account's orders
                                        # - necessary to track if any rawOrders are for an unknown or invalid account
                                        remaining_orders = [order for order in remaining_orders
                                                            if order not in account_orders]

                                        # (2) review raw EXIT orders and place as appropriate
                                        account.exit_results.extend(self.place_new_exits(account_orders, broker,
                                                                                         trade_register, account))

                                        # (3) review raw ENTRY orders and place as appropriate
                                        account.entry_results.extend(self.place_new_entries(account_orders,
                                                                                            broker, account))

                                    else:
                                        # An error was returned when finding a valid broker/trade register. These orders
                                        # cannot be placed.  The error will have been logged and reported earlier when
                                        # the error arose.
                                        msg = f'Could not establish a broker connection for {account.alias} ' \
                                              f'(ID: {account.id}). {len(account_orders)} ' \
                                              f'{self.plo(account_orders)} will not be placed'
                                        account.log.info(msg)
                                        account.runtime_errors.append(msg)

                            # Did we successfully process all the orders from the Order File provided?
                            if remaining_orders:
                                # No - there are orders left over which were not placed in an account
                                for order in remaining_orders:
                                    if order.account_id not in [a.id for a in self.trading_accounts]:
                                        order.set_status('INVALID', 'ERROR',
                                                         f'Invalid account ID {order.account_id}')
                                    else:
                                        order.set_status('INVALID', 'ERROR',
                                                         f'Invalid strategy {order.qa_strategy_id} '
                                                         f'for account ID {order.account_id}')

                        # For each account, email the Order Placement log to the user.  Note this is
                        # different to the Daily Report.  No position data is reported in the Order Placement log.
                        # TODO: refactor this - move it into a separate method
                        if orders_found:
                            for account in self.trading_accounts:

                                self.order_processing_status = f'Sending email log report for {account.alias} orders'

                                if account.ib_type == 'FUND ADVISORY' and not self.config['Report on Advisor Account']:
                                    continue

                                LOG.info(f'Sending email log report for {account.alias}.')

                                # retrieve account log file
                                account_log_filepath = [h for h in account.log.handlers
                                                        if type(h) == logging.FileHandler][0].baseFilename

                                # build email body
                                all_orders = account.entry_results + account.exit_results

                                session_date = US_MARKET.next_session.trading_date
                                if all_orders:
                                    order_dates = list(set([ordr.date for ordr in all_orders]))
                                    if order_dates:
                                        session_date = order_dates[0]

                                strategy_ids = [s.qa_id for s in self.strategies
                                                if s.qa_id in account.strategy_allocations.keys() and
                                                account.strategy_allocations[s.qa_id] > 0]
                                strategy_ids.sort()

                                email_body = (f'{account.alias} ({account.id}) ORDER PLACEMENT REPORT for '
                                              f'{session_date} <br><br>')

                                if strategy_ids:
                                    for s in strategy_ids:
                                        strategy_entries = [ordr for ordr in account.entry_results
                                                            if ordr.qa_strategy_id == s
                                                            and not ordr.is_exit]
                                        strategy_exits = [ordr for ordr in all_orders
                                                          if ordr.qa_strategy_id == s and ordr.is_exit]
                                        strategy_all = strategy_entries + strategy_exits
                                        errors_text = ''
                                        for order in strategy_all:
                                            for e in order.errors:
                                                errors_text += '<br>\t• ' + order.ticker + ': ' + e
                                        errors_text = errors_text if errors_text else '<br>\tNo errors reported.'
                                        entries_text = '<br>'.join(
                                            [f'\t{ordr.action} {ordr.associated_trade.order_size} {ordr.ticker} '
                                             f'{ordr.final_instrument_type} @ '
                                             f'{ordr.order_type} {ordr.order_price if ordr.order_type == "LMT" else ""}'
                                             for ordr in strategy_entries
                                             if ordr.status in ['PLACED', 'OPEN', 'COMPLETE']])
                                        if not strategy_entries:
                                            entries_text = 'None'
                                        else:
                                            print('debug')
                                        exits_text = '<br>'.join(
                                            [f'\t{ordr.action} {ordr.associated_trade.order_size} {ordr.ticker} '
                                             f'{ordr.final_instrument_type} @ '
                                             f'{ordr.order_type} {ordr.order_price if ordr.order_type == "LMT" else ""}'
                                             for ordr in strategy_exits
                                             if ordr.status in ['PLACED', 'OPEN', 'COMPLETE']])
                                        if not strategy_exits:
                                            exits_text = 'None'

                                        email_body += f'{"#" * (41 + len(s))}<br>' \
                                                      f'{"#" * 10}   <strong>STRATEGY: {s}   </strong>{"#" * 15}<br>' \
                                                      f'{"#" * (41 + len(s))}<br><br>' \
                                                      f'<strong>ERRORS:</strong> <br>{errors_text}<br><br>' \
                                                      f'<strong>ENTRY ORDERS PLACED:</strong> ' \
                                                      f'<br>{entries_text}<br><br>' \
                                                      f'<strong>EXIT ORDERS PLACED:</strong> <br>{exits_text}<br><br>'
                                else:
                                    email_body += f'<br><br>No entry or exit orders were placed.'

                                # send email
                                EMAIL_MANAGER.send_email(subject=f'{account.alias} Order Placement Log',
                                                         body=email_body,
                                                         text_type='html',
                                                         attachments_file_paths=[])
                                LOG.info(f'Log report for {account.alias} emailed.\n')

                            # Once all accounts have been processed, notify the user to add another order file
                            LOG.info(f'Waiting for order file. Please drag or save an order file to '
                                     f'{ADDR.folder_order_files} to initiate trading.\n')

                        self.order_processing_status = ''

                    # after processing a set of orders, sleep for a period
                    if OMS_SETTINGS.process_orders_basis == 'SCHEDULED':
                        # if we are not in an active scheduled window, sleep till the window opens
                        while not self.stop_signal() and not self.is_time_to_process_orders() and \
                                not self.manual_order_process_button_clicked:
                            sleeper.sleep(1)
                    else:
                        # if we are waiting for manual order button to be pressed, sleep 1 second.
                        sleeper.sleep(1)

                sleeper.sleep(1)

        except Exception as exc:
            # captures unhandled exceptions
            self.raise_oms_error((1, exc, True))

    @staticmethod
    def order_error_string(order):
        # used for order placement log; receives an order object, and parses, to generate a simplified string to
        # report in email
        errors_text = ''

        if order.errors:
            errors_text += '<br>\t' + order.ticker + ': ' + \
                           ','.join([e for e in order.errors])

        return errors_text

    @property
    def scheduled_order_time(self):

        yesterday = datetime.combine((US_MARKET.current_time - timedelta(days=1)).date(),
                                     OMS_SETTINGS.scheduled_order_processing_time).replace(tzinfo=US_MARKET.zone_info)
        yesterday_1 = yesterday + timedelta(hours=1)
        yesterday_2 = yesterday + timedelta(hours=2)

        today = datetime.combine(US_MARKET.current_time.date(),
                                 OMS_SETTINGS.scheduled_order_processing_time).replace(tzinfo=US_MARKET.zone_info)
        today_1 = today + timedelta(hours=1)
        today_2 = today + timedelta(hours=2)

        tomorrow = datetime.combine((US_MARKET.current_time + timedelta(days=1)).date(),
                                    OMS_SETTINGS.scheduled_order_processing_time).replace(tzinfo=US_MARKET.zone_info)
        tomorrow_1 = tomorrow + timedelta(hours=1)
        tomorrow_2 = tomorrow + timedelta(hours=2)

        valid_sots = [t for t in [yesterday, yesterday_1, yesterday_2,
                                  today, today_1, today_2,
                                  tomorrow, tomorrow_1, tomorrow_2]
                      if t + timedelta(minutes=1) > US_MARKET.current_time]

        return min(valid_sots) if valid_sots else US_MARKET.current_time + timedelta(days=1)

    def is_time_to_process_orders(self):

        if OMS_SETTINGS.process_orders_basis == 'SCHEDULED' and self.initialised:
            if US_MARKET.current_time.replace(second=0, microsecond=0) == self.scheduled_order_time:
                # orders will be placed within the 1 minute window of the designated time.
                return True

        return False

    def retrieve_broker(self, account, refresh_data=True):

        # A 'broker' in the OMS is an instance of the BrokerApp class. It provides a shell around a separate instance
        # of the IB_API class, and functions as an IB API controller.

        # A single broker instance can manage multiple brokerage accounts, in keeping with IB's own structure, where
        # a Fund Advisory account can manage multiple sub-accounts.

        # Stand-Alone Accounts - have a unique client ID
        # Fund Advisory Accounts - have a unique client ID, which must match the TWS Master Client ID
        # Managed Accounts - do not require a client ID, as they are manged through the Fund Advisory account broker.

        # For all account activities (reconciliation, orders, reporting) the correct broker account must be used.
        # Before any account activities, the OMS will retrieve the correct broker connection and trade register.  Each
        # time the broker is retrieved, it is validated to ensure that the connection is live and for the correct
        # account and when refresh_data == True the latest data from the broker is obtained.
        with self._broker_lock:
            LOG.info(f'Getting broker and account for {account.id}')
            broker = [b for b in self.brokers if account.id in b.api.connected_accounts]

            # retrieve or create the broker and trade register objects
            if broker and not self.stop_signal():
                broker = broker[0]
                try:
                    trade_register = [tr for tr in broker.api.trade_registers if tr.account_id == account.id][0]
                except IndexError:
                    trade_register = TradeRegister(account.id)
                    broker.api.trade_registers.append(trade_register)
            else:
                # a TWS/IBG connection for this account has not yet been made - make one.
                broker = BrokerApp(account.id, account.alias, port=account.ib_port,
                                   client_id=account.ib_client_id)
                trade_register = TradeRegister(account.id)
                broker.api.trade_registers.append(trade_register)

            if account.ib_type == 'FUND ADVISORY':
                # add child trade registers to broker connection, so that all callbacks can be mapped
                child_accounts = [a for a in self.trading_accounts
                                  if a.ib_type == 'MANAGED'
                                  and a.parent_account == account.id
                                  and a.id not in [t.account_id for t in broker.api.trade_registers]]
                for child_account in child_accounts:
                    try:
                        ctr = [tr for tr in broker.api.trade_registers if tr.account_id == child_account.id][0]
                    except IndexError:
                        ctr = TradeRegister(child_account.id)
                    broker.api.trade_registers.append(ctr)

            connected, msg, new_msg = False, '', ''
            # connect to the broker (or verify existing connection) and get the most recent broker updates
            while not connected and not self.stop_signal():
                try:
                    if broker.api.isConnected():
                        broker.api.verify_connection()
                    else:
                        LOG.debug('Not connected. Attempting connection')
                        broker.connect_to_api_account()
                        LOG.debug('Connection successful')
                    if broker and broker not in self.brokers:
                        self.brokers.append(broker)
                    # next 2 lines redundantly reloads managed accounts data if the parent account has just been loaded;
                    # might be a better/quicker way of doing it, but there are instances where the child account will be
                    # connected without re-calling the parent account, so this is necessary for the moment.
                    broker.update_current_prices(account.id, refresh=refresh_data)
                    broker.update_trade_register_from_api(refresh=refresh_data)
                    return broker, trade_register
                except (ConnectionError, ValueError, TimeoutError) as error_info:
                    if 'Could not connect to IB account' in error_info.args[0][1]:
                        new_msg = (f'There is a problem connecting to {account.alias} ({account.id}) via TWS/IBG '
                                   f'port {account.ib_port} - is TWS running?  Or review account and port settings in '
                                   f'config file? Or accept paper account warning? Error message:  '
                                   f'{error_info.args[0][1]}')
                        if new_msg != msg:
                            account.log.error(new_msg)
                    elif 'Invalid Account ID' in error_info.args[0][1]:
                        # deprecated
                        new_msg = (f'Error accessing account {account.alias} ({account.id}) via port '
                                   f'{account.ib_port} - review account ID and port settings in config file.  This '
                                   f'error can also be thrown by Paper accounts where the Paper account warning has '
                                   f'not been accepted yet.')
                        if new_msg != msg:
                            account.log.error(new_msg)
                    elif 'callback not received from Account' in error_info.args[0][1]:
                        new_msg = error_info.args[0][1]
                        if new_msg != msg:
                            account.log.warning(new_msg)
                    else:
                        self.raise_oms_error((2,
                                              ConnectionError(f'Unhandled error connecting to {account.alias} '
                                                              f'({account.id}) '
                                                              f'Error Message: {error_info.args[0]}. '
                                                              f'Please address the error and click the Restart '
                                                              f'OMS button'),
                                              True))
                    msg = new_msg
                    connected, new_msg = False, ''
                    sleeper.sleep(1)

    def reconcile(self, broker, account, close_orphan_positions=False):

        # Before new orders are placed the accounts Trade Register is reconciled with the broker data.

        LOG.info(f'{account.alias}: Reconciling broker positions with trade register.')

        # Identify any unreconciled (1) BROKER POSITIONS or (2) TRADES
        excess_positions, excess_trades = broker.reconcile_trades_and_broker(account.id)

        # (1) review any unreconciled BROKER POSITIONS - i.e. there is a position at the broker which is not also
        # found in the trade register.
        for contract, qty in excess_positions:

            ignored_qty = sum([x[1] for x in account.ignored_positions if x[0] == contract.conId])

            if qty == ignored_qty:
                # this position can be ignored, according to the config file.
                pass
            else:
                # this is an orphaned position
                qty = qty - ignored_qty
                if close_orphan_positions:
                    # close the orphaned position and notify user
                    action = 'BUY' if qty < 0 else 'SELL'
                    closing_order = o.market_order(action=action, quantity=abs(qty),
                                                   order_ref='close_orphan', tif='DAY')
                    closing_order.account = account.id
                    broker.place_order(contract, closing_order)
                    account.log.error(f"{account.alias} Position of {qty} x {contract.symbol} @ "
                                      f"{contract.exchange} could not be reconciled with trade records. "
                                      f"Position has been closed.")
                else:
                    # don't close the orphaned position - just notify user
                    # currently, this is default behaviour
                    account.log.error(f"{account.alias} Position of {qty} x {contract.symbol} @ "
                                      f"{contract.exchange} could not be reconciled with trade records. No action "
                                      f"has been taken to correct this.")

        # (2) review any unreconciled TRADES - i.e. trade records in our trade register which indicate a position
        # which is not also found at the broker.
        for t in excess_trades:
            # No action currently taken - just notify user
            account.log.error(f"Account {account.id} Trade Register shows an open trade for "
                              f"{t.current_position} x {t.contract.symbol} at {t.contract.exchange} which is not "
                              f"matched at the broker.  No action has been taken to correct this.")

        # (3) review duplicated open trades
        trade_register = broker.api.retrieve_trade_register(account.id)
        open_trades = [t for t in trade_register.trades if t.status in ['PLACED', 'OPEN']]
        duplicated_trades_list = [(t.ticker, t.strategy.qa_strategy_model) for t in open_trades]
        duplicated_trades_list = list(set([x for x in duplicated_trades_list if duplicated_trades_list.count(x) > 1]))
        for dt in duplicated_trades_list:
            ticker = dt[0]
            strat_model = dt[1]
            self.raise_oms_error((1, ValueError(f'Duplicated open trades found in {ticker} for {strat_model}. '
                                                f'Contact the developer.'),
                                  True))
            dup_trades = [t for t in open_trades if t.ticker == ticker and t.strategy.qa_strategy_model == strat_model]
            account.log.error(f'Duplicated open trades found in {ticker} for {strat_model}:')
            for t in dup_trades:
                print(f' - Trade {t.trade_id}: Order for {t.order_size} {t.ticker} shares, current position = '
                      f'{t.current_position if t.current_position else 0}, '
                      f'trade record created {datetime.strftime(t.created_datetime, "%Y-%m-%d %H:%M:%S")} market time')
            print("Only one trade can remain. Enter the trade ID that should be KEPT.  "
                  "(The other trades will be closed, orders cancelled and deleted from the register.)")
            keep_trade = ''
            while keep_trade not in [t.trade_id for t in dup_trades]:
                keep_trade = input('Enter the trade # TO BE KEPT: ')
                try:
                    keep_trade = int(keep_trade)
                except ValueError:
                    pass
            for t in [t for t in dup_trades if t.trade_id != keep_trade]:
                account.log.info(f'Trade {t.trade_id}: closing trade and deleting from register')
                if t.current_position != 0:
                    order = o.market_order({'SELL' if t.strategy.direction == 'LONG' else 'BUY'},
                                           t.current_position, f'Cancel_Trade_{t.trade_id}', 'GTC')
                    o.account = account.id
                    broker.place_order(t.contract, order)
                    account.log.info(f'Market order placed to close {t.ticker} position of {t.current_position}.')
                for eor in [t.entry_order_record] + [e for e in t.exit_order_records]:
                    if eor.is_active_at_broker(broker):
                        broker.cancel_order(eor.order.orderId)
                        account.log.info(f'Open Order {eor.order.orderId} ({eor.order.orderRef}) cancelled')
                trade_register.trades.remove(t)
                trade_register.save_trade_list()
                account.log.info(f'Trade {t.trade_id} deleted from Trade Register')

        if excess_positions or excess_trades or duplicated_trades_list:
            LOG.info(f'{account.alias}: Reconciliation complete.\n')
        else:
            LOG.info(f'{account.alias}: Reconciliation complete.  No orphaned positions, orphaned trades or '
                     f'duplicated trades found.\n')

    def place_new_exits(self, raw_orders, broker, trade_register, account):

        # from the imported rawOrders for this account, identify those that are exit orders for today
        todays_exit_orders = [order for order in raw_orders if order.is_exit]

        if todays_exit_orders:
            todays_exit_orders.sort(key=lambda x: x.order_rank)
            LOG.info(f'{account.alias} EXIT ORDERS: {len(todays_exit_orders)} exit {self.plo(todays_exit_orders)} for '
                     f'account {account.alias} were found in the imported file.\n')
        else:
            LOG.info(f'{account.alias} EXIT ORDERS: No exit orders for {account.alias} '
                     f'were found in the imported file.\n')

        # process each order one by one
        for order in todays_exit_orders:

            self.stop_signal()

            # check order is valid
            valid, reasons = order.valid(valid_accounts=self.trading_accounts, valid_strategies=self.strategies,
                                         market=US_MARKET)
            if not valid:
                order.set_status('INVALID', account.log.error, ';'.join(reasons))
                continue

            # uniquely identify the trade in the trade register that this exit order belongs to
            t = [trd for trd in trade_register.trades if
                 trd.active and
                 trd.strategy.qa_strategy_model == order.qa_strategy_model and
                 trd.ticker == order.ticker]

            if len(t) > 1:
                # multiple trades match the order data - error
                order.set_status('INVALID', account.log.error,
                                 f'Multiple trades found matching {order.summary_text}. Exit not placed')
            elif not t:
                # there's no matching trade in the trade register. e.g. RT thinks there is a position, but there's not
                order.set_status('REJECTED', account.log.warning,
                                 f'No matching trade found for exit order {order.action} {order.ticker}. '
                                 f'Exit not placed')
            else:
                # the trade that this exit order applies to has been uniquely identified
                t = t[0]
                order.associated_trade = t

                # confirm that the trade has an open position
                if t.active and t.current_position != 0:
                    exit_order = None
                    order_ref = OrderRef(order.qa_strategy_model, t.trade_id, f'EXIT-{t.next_exit_id(broker)}').ref

                    # confirm that the order is not already at the market
                    active_exits = [e for e in t.exit_order_records
                                    if e.status(broker, account.id) in ['Filled', 'Active']]
                    if active_exits:
                        nl = '\n • '
                        exits_list = ' • ' + nl.join([f'{e.order.action} x {e.order.totalQuantity}, order ref '
                                                      f'{e.order.orderRef},  created '
                                                      f'{datetime.strftime(e.created_datetime, "%Y-%m-%d %H:%M:%S")},'
                                                      f'status = {e.status(broker, account.id)}'
                                                      for e in active_exits])
                        order.set_status('REJECTED', account.log.info,
                                         f'Trade {t.trade_id} already has active exit order(s): {exits_list}\n'
                                         f'New exit not placed.')

                    # what order type does this exit order use?
                    if order.status != 'REJECTED':
                        if order.order_type == 'LMT':

                            # conform the exit order price to IB min tick requirements
                            contract_details = broker.get_contract_details(t.contract)
                            if contract_details:
                                contract_details = contract_details[0]
                            try:
                                _order_price = broker.conform_price_to_broker_min_tick(order.order_price,
                                                                                       contract_details.marketRuleIds)
                            except ValueError:
                                account.log.info(f'Could not find min tick price from TWS for {order.ticker}. Price '
                                                 f' {order.order_price} will be rounded to 2 decimal places.')
                                _order_price = round(order.order_price, 2)

                            if _order_price != order.order_price:
                                order.set_status('PENDING', account.log.warning,
                                                 f'Exit order price for {order.ticker} adjusted from '
                                                 f'{order.order_price} to {_order_price} to conform to IB '
                                                 f'min tick requirements.')

                            # create the Limit Exit Order object
                            exit_order = o.limit_order(action=order.action, quantity=abs(t.current_position),
                                                       limit_price=_order_price, order_ref=order_ref, ORTH=False,
                                                       tif=order.tif)

                        elif order.order_type == 'MKT':

                            # create the Market Exit Order object
                            exit_order = o.market_order(action=order.action, quantity=abs(t.current_position),
                                                        order_ref=order_ref, tif=order.tif)

                        else:

                            # Currently the code only handles 'LMT' and 'MKT' exit orders. This will be updated as
                            # more exit orders are required.
                            order.set_status('INVALID', account.log.error,
                                             f'Unknown order type {order.order_type} for {order.summary_text}. '
                                             f'Exit not placed')

                    if exit_order:

                        # create the necessary records for the order
                        eor = o.OrderRecord(exit_order)
                        account.log.info(f'Placing Exit Order {order.action} {abs(t.current_position)} x '
                                         f'{order.ticker} {t.contract.secType} via {t.contract.exchange}')

                        # place the order at the broker
                        exit_order.account = account.id
                        order_time = datetime.now()
                        order_id = t.place_exit_order(eor, broker)

                        # wait for IB callbacks, IB errors, or a timeout
                        # Background: when an IB callback is received it is saved immediately to the trade object. So
                        # the code monitors for order activity by monitoring the trade variables.
                        timer = util.CodeTimer()
                        while (not eor.orderState
                                and not [e for e in t.errors
                                         if e.received > order_time
                                         and e.code not in self.config['Ignore IB Errors']]
                                and timer.end() < 10):
                            sleeper.sleep(0.1)
                        LOG.debug(f'Order {order_id} - Status = {eor.status(broker, account.id)}')
                        new_errors = [e for e in t.errors
                                      if e.received > order_time
                                      and e.code not in self.config['Ignore IB Errors']]

                        order.ib_errors = [e for e in t.errors if e.received > order_time]

                        # did we receive an IB callback or error?
                        if eor.orderState or new_errors:
                            # the exit order has been processed by the broker

                            # allow time for secondary callbacks & errors to be received
                            sleeper.sleep(2)
                            LOG.debug(f'Order {order_id} - Status = {eor.status(broker, account.id)}')
                            msg = ', '.join([f'{err.code}: {err.string}' for err in t.errors
                                             if err.received > order_time])

                            # log outcome of placing exit order
                            if eor.status(broker, account.id) in ['Active', 'Partial Fill', 'Filled']:

                                # exit order was successfully placed
                                order.set_status('PLACED', account.log.info, f'Exit Order successfully placed. {msg}')

                            elif t.status == 'CANCELLED':

                                # order has been cancelled/rejected because of an error from IB.  When an error is
                                # received from IB for an exit order, and the entry order has not been filled, the
                                # whole trade is cancelled. This will be an unusual/exceptional circumstance.
                                # See the trades.TradeRecord.process_ib_error function for trade error handling.
                                order.set_status('REJECTED', account.log.warning,
                                                 f'Exit Order: {order.action} {t.current_position} x '
                                                 f'{order.ticker} cancelled/rejected by '
                                                 f'broker. Trade has been cancelled: {msg}')

                            else:
                                # the order was cancelled, but the trade remains active at the broker.
                                order.set_status('REJECTED', account.log.warning,
                                                 f'Exit Order rejected by broker. Trade remains active. Broker '
                                                 f'errors: {msg}')

                        else:
                            # no callback received from broker
                            broker.cancel_order(order_id)
                            order.set_status('REJECTED', account.log.warning,
                                             f'Error placing exit order for {t.current_position} x {order.ticker} '
                                             f'via {t.contract.exchange} - no callback received. Exit not placed.')
                else:
                    # the exit order is for a position that doesn't exist
                    order.set_status('REJECTED', account.log.warning,
                                     f'Exit Order {order.action} {order.ticker} is for a trade that has no current '
                                     f'position. The Exit Order will not be placed.')

            SIGNALS.raw_order_update.emit()

        LOG.info("Finished processing exit orders\n")

        return todays_exit_orders

    def place_new_entries(self, account_orders, broker, account):

        # from the imported rawOrders for this account, identify those that are entry orders for today
        todays_entry_orders = [order for order in account_orders if order.is_entry]

        if not self.stop_signal():
            if todays_entry_orders:
                account.log.info(f'{account.alias} ENTRY ORDERS - {len(todays_entry_orders)} entry '
                                 f'{self.plo(todays_entry_orders)} for {account.alias} found in the imported file.\n')
            else:
                account.log.info(f'{account.alias} ENTRY ORDERS - no entry orders for {account.alias} were found in the '
                                 f'imported file.\n')
                return todays_entry_orders
        else:
            return todays_entry_orders
        scheduled_entry_date = todays_entry_orders[0].date

        # get the strategy ids found in today's orders
        todays_strategy_ids = list(set([order.qa_strategy_id for order in todays_entry_orders]))
        trade_register = broker.api.retrieve_trade_register(account.id)
        n = 0

        account_nlv = broker.nlv(account.id, 'USD', refresh=True)

        # Entry orders are grouped by strategy before processing in strategy batches.  This is not essential but
        # allows simpler tracking of max orders/max positions.
        for qa_strategy_id in todays_strategy_ids:

            # retrieve the strategy specifications, as loaded from the config file
            strategy = [s for s in self.strategies if s.qa_id == qa_strategy_id]
            if strategy:
                strategy = strategy[0]
            else:
                account.log.error(f'Unknown strategy {qa_strategy_id} found in orders file.  No orders for this '
                                  f'strategy will be placed.  Please add the strategy to the config file.\n')
                for order in [order for order in account_orders if
                              order.qa_strategy_id == qa_strategy_id and order.is_entry]:
                    order.set_status('INVALID', account.log.warning, 'Unknown strategy')
                continue

            LOG.info(f'PROCESSING ORDERS FOR STRATEGY {strategy.qa_id}\n')

            # how many open positions does this strategy already have?
            strategy_open_positions = [t for t in trade_register.trades if
                                       t.status in ['PLACED', 'OPEN'] and
                                       t.strategy.qa_id == qa_strategy_id
                                       and not [eor for eor in t.exit_order_records
                                                if eor.order.orderType == 'MKT' and
                                                eor.is_active_at_broker(broker, account.id)]]
            strategy.open_exposure = sum([t.exposure for t in strategy_open_positions if t.status == 'OPEN'])
            strategy.todays_exposure = sum([t.exposure for t in strategy_open_positions
                                            if t.status == 'PLACED'
                                            and t.scheduled_entry_order_date == scheduled_entry_date])

            # retrieve today's orders for this strategy, and sort them by ascending order rank (i.e. 1 is first)
            strategy_todays_orders = [order for order in account_orders if
                                      order.qa_strategy_id == qa_strategy_id and order.is_entry]
            strategy_todays_orders.sort(key=lambda x: x.order_rank)

            # evaluate and place each order in the list of strategy orders
            for order in strategy_todays_orders:

                self.stop_signal()

                # setup
                n += 1
                account.log.info(f'{account.alias} entry order {n} of {len(todays_entry_orders)} - '
                                 f'{order.summary_text}')

                # check the order is valid
                valid, reasons = order.valid(valid_accounts=self.trading_accounts, valid_strategies=self.strategies,
                                             market=US_MARKET)
                if not valid:
                    order.set_status('INVALID', account.log.error, ';'.join(reasons))

                # check that the same order is not already active at the broker
                if account.id in broker.api.raw_open_orders.keys():
                    ticker_entry_orders = [order_ref for order_ref, order_details in
                                           broker.api.raw_open_orders[account.id].items()
                                           if order_details[1].symbol == order.ticker
                                           and 'ENTRY' in order_ref
                                           and order.qa_strategy_model in order_ref]
                    if len(ticker_entry_orders) == 1:
                        order.set_status('REJECTED', account.log.info,
                                         f'ENTRY order {ticker_entry_orders[0]} is already active at the broker')
                    elif len(ticker_entry_orders) > 1:
                        order.set_status('REJECTED', account.log.error,
                                         f'Multiple entry orders found for {strategy.qa_id} for '
                                         f'{order.ticker}.  New order not placed. Review multiple orders for errors')

                # check that the strategy does not already have an open trade for this ticker
                # open trades with an active market order are excluded - i.e. the trade will exit at the open, so a
                #  new trade (the current order) can still be validly placed without duplicating positions.
                # NOTE: this functionality will also prevent scaling into positions with multiple entry orders.
                #       i.e. the OMS does not currently support scaling in.
                strategy_model_open_positions = [t for t in trade_register.trades if
                                                 t.status in ['PLACED', 'OPEN'] and
                                                 t.strategy.qa_strategy_model == order.qa_strategy_model
                                                 and not [eor for eor in t.exit_order_records
                                                          if eor.order.orderType == 'MKT' and
                                                          eor.is_active_at_broker(broker, account.id)]]
                ticker_strategy_positions = [t for t in strategy_model_open_positions if
                                             t.ticker == order.ticker]
                if ticker_strategy_positions:
                    for t in ticker_strategy_positions:
                        t.calculate_status()
                        if t.status == 'OPEN':
                            order.set_status('REJECTED', account.log.info,
                                             f'Strategy {order.qa_strategy_model} already has open position '
                                             f'of {t.current_position} at broker for '
                                             f'Trade {ticker_strategy_positions[0].trade_id} for {order.ticker}')
                        elif t.status in ['DRAFT', 'PLACED']:
                            if t.entry_order_record.is_active_at_broker(broker, account.id):
                                order.set_status('REJECTED', account.log.info,
                                                 f'Strategy {order.qa_strategy_model} already has a draft order Id '
                                                 f'{t.entry_order_record.orderId} in '
                                                 f'Trade {ticker_strategy_positions[0].trade_id} for {order.ticker}')
                        else:
                            order.set_status('REJECTED', account.log.info,
                                             f'Strategy {order.qa_strategy_model} already active at broker with'
                                             f' Trade {ticker_strategy_positions[0].trade_id} for {order.ticker} - '
                                             f'status could not be determined')

                if order.status in ['REJECTED', 'INVALID']:
                    account.log.info(f'Order {n}: Order could not be placed. Order processing complete.')
                    continue

                # attempt to place entry order for each security type until an order is successfully placed
                attempt_second_security_type = True  # set to false for certain errors below
                security_types = account.strategy_security_types[strategy.qa_id]
                for security_type in security_types:

                    if security_type in [None, ''] or not attempt_second_security_type:
                        continue

                    order.status = 'PENDING'

                    # 1. notify user if we have moved on to the secondary security type
                    if security_type != security_types[0]:
                        account.log.info(f'Order {n}: trying strategy secondary security type ({security_type})')

                    # 2. get the contract details for this security from the broker
                    contract_details = None
                    contract = broker.get_contract_from_ticker(ticker=order.ticker,
                                                               currency=US_MARKET.currency,
                                                               exchange='SMART',
                                                               secType=security_type)
                    if contract:
                        contract_details = broker.get_contract_details(contract)
                        if contract_details:
                            contract_details = contract_details[0]
                    if not contract or not contract_details:
                        order.set_status('REJECTED', account.log.warning, f'[{security_type}] Could not identify an '
                                                                          f'IB contract from: '
                                                                          f'{order.ticker} {security_type}, '
                                                                          f'{US_MARKET.currency}.')
                    else:
                        # 2A. Check that order date is for a valid session for the contract
                        contract_session = contract_details.liquidHours.split(';')
                        contract_session_dates = [datetime.strptime(s[:8], '%Y%m%d').date() for s in contract_session]
                        if order.date not in contract_session_dates:
                            order.set_status('REJECTED', account.log.error,
                                             f'Order date ({datetime.strftime(order.date, "%Y-%m-%d")}) is a non '
                                             f'market date - is it a public holiday?')

                        # 2B. for short orders, check margin rates and borrow costs
                        if order.direction == 'SHORT':

                            reject_reason, short_margin_value, borrow_fee_rate = [], None, None

                            # retrieve shortMargin and FeeRate
                            con_id = float(contract.conId)
                            if contract.secType == 'CFD':
                                con_id = float(contract_details.underConId)

                            # check short margin
                            if not self.ib_stock_margin_data.empty:
                                short_margin_df = self.ib_stock_margin_data[self.ib_stock_margin_data['CON'] == con_id]
                                if not short_margin_df.empty:
                                    short_margin_value = short_margin_df.iloc[
                                        0, short_margin_df.columns.get_loc('ShortMargin')]
                                    if type(short_margin_value) not in [float, int]:
                                        try:
                                            short_margin_value = float(short_margin_value)
                                        except (TypeError, ValueError):
                                            # can be blank or 'Default'
                                            short_margin_value = None
                                    if short_margin_value and short_margin_value > self.config['Max Short Margin']:
                                        reject_reason.append(f'Short Margin value of {short_margin_value} exceeds '
                                                             f'maximum of {self.config["Max Short Margin"]}')
                            else:
                                # v 0.1.16.7 - only report an error if no file from IB obtained
                                account.log.warning(f'Could not obtain ShortMargin value for {order.ticker}. Order '
                                                    f'will be processed without doing a ShortMargin check.')
                                order.errors.append(f'Could not obtain ShortMargin value. Order will be '
                                                    f'processed without doing a ShortMargin check.')

                            # check fee rate
                            if not self.ib_stock_borrow_costs.empty:
                                borrow_costs_df = self.ib_stock_borrow_costs[self.ib_stock_borrow_costs['CON'] ==
                                                                             con_id]
                                if not borrow_costs_df.empty:
                                    borrow_fee_rate = borrow_costs_df.iloc[
                                        0, borrow_costs_df.columns.get_loc('FEERATE')]
                                    if type(borrow_fee_rate) not in [float, int]:
                                        try:
                                            borrow_fee_rate = float(borrow_fee_rate)
                                        except (TypeError, ValueError):
                                            borrow_fee_rate = None
                                    if borrow_fee_rate and borrow_fee_rate > self.config['Max Fee Rate']:
                                        reject_reason.append(f'Fee Rate of {borrow_fee_rate} exceeds maximum '
                                                             f'of {self.config["Max Fee Rate"]}')
                            else:
                                # v 0.1.16.7 - only report an error if no file from IB obtained
                                account.log.warning(f'Could not obtain FEERATE value for {order.ticker}. Order will be'
                                                    f' placed without doing a FEERATE check.')
                                order.errors.append(f'Could not obtain FEERATE value. Order will be '
                                                    f'processed without doing a FEERATE check.')

                            # reject or approve
                            if reject_reason:
                                order.set_status('REJECTED', account.log.warning, ', '.join(reject_reason))
                            else:
                                account.log.info(f'{contract.symbol}: Instrument passes margin & fee rate checks - '
                                                 f'Short Margin ({short_margin_value}) less '
                                                 f'than maximum permitted ({self.config["Max Short Margin"]}), '
                                                 f'Fee Rate ({borrow_fee_rate}) less than maximum permitted '
                                                 f'({self.config["Max Fee Rate"]})')

                    if order.status != 'REJECTED':
                        # 3. adjust the order price to match the min_tick from the broker
                        if order.order_type == 'LMT':
                            _order_price = broker.conform_price_to_broker_min_tick(order.order_price,
                                                                                   contract_details.marketRuleIds)
                            if _order_price != order.order_price:
                                order.set_status(order.status, account.log.warning,
                                                 f'[{security_type}] Order price adjusted from {order.order_price} to '
                                                 f'{_order_price} to conform to IB min tick requirements.')
                        else:
                            _order_price = Decimal(str(order.prior_close))

                        # 4. Calculate position size for order from strategy and order data
                        pos_size = int((Decimal(str(order.pos_size_pct * account.strategy_allocations[strategy.qa_id]))
                                        * account_nlv) / _order_price)
                        if order.pos_size_pct == 0:
                            order.set_status('REJECTED', account.log.warning,
                                             f'[{security_type}] Order specifies a zero % position size.')
                        elif account.strategy_allocations[strategy.qa_id] == 0:
                            order.set_status('REJECTED', account.log.warning,
                                             f'[{security_type}] Account strategy allocation is zero %.')
                        elif pos_size <= 0:
                            order_value = (Decimal(
                                str(order.pos_size_pct * account.strategy_allocations[strategy.qa_id])) * account_nlv)
                            if order_value < _order_price:
                                attempt_second_security_type = False
                                order.set_status('REJECTED', account.log.warning,
                                                 f'[{security_type}] The total calculated order value based on '
                                                 f'allocations'
                                                 f' is less than the instrument order price. '
                                                 f' No entry order will be placed. '
                                                 f'[Calculated order value = ${round(order_value, 2):.2f}, '
                                                 f'calculated from PosSize%={round(order.pos_size_pct * 100, 2)}%, '
                                                 f'AccountStratAllocation%='
                                                 f'{account.strategy_allocations[strategy.qa_id]*100}%, '
                                                 f'AccountNLV=${account_nlv}, OrderPrice=${_order_price:.2f}]. '
                                                 f'Secondary security type will not be tried.')
                            else:
                                attempt_second_security_type = False
                                order.set_status('REJECTED', account.log.warning,
                                                 f'[{security_type}] Account funds (${account_nlv}) are not sufficient '
                                                 f'to place order. '
                                                 f'[Calculated position size = {pos_size}, '
                                                 f'order value = ${round(pos_size * _order_price,2)}, '
                                                 f'calculated from PosSize%= {round(order.pos_size_pct, 2)}, '
                                                 f'AccountStratAllocation={account.strategy_allocations[strategy.qa_id]}, '
                                                 f'OrderPrice={_order_price}].  Secondary security type will not be '
                                                 f'tried.')

                        # 5. Check that the order will not exceed exposure limits

                        new_exposure_from_this_order = None
                        if order.status != 'REJECTED':
                            new_exposure_from_this_order = ((_order_price * pos_size) /
                                                            (account_nlv *
                                                             Decimal(
                                                                 str(account.strategy_allocations[strategy.qa_id]))))

                            valid_exposure = (strategy.open_exposure +
                                              strategy.todays_exposure +
                                              new_exposure_from_this_order) < order.max_exposure
                            valid_daily_exposure = (strategy.todays_exposure +
                                                    new_exposure_from_this_order) < order.max_new_exposure

                            account.log.info(f'[{security_type}] - {contract.symbol} - calculating exposure.\n'
                                             f'Order Max Exposure: {order.max_exposure}\n'
                                             f'Order Max New Exposure: {order.max_new_exposure}\n'
                                             f'Strategy Open Exposure: {round(strategy.open_exposure, 5)}\n'
                                             f'Strategy Today Exposure: {round(strategy.todays_exposure, 5)}\n'
                                             f'New Exposure from this order: {round(new_exposure_from_this_order, 5)}\n'
                                             f'Valid Exposure: {valid_exposure}\n'
                                             f'Valid Daily Exposure: {valid_daily_exposure}')

                            if order.status != 'REJECTED' and not valid_exposure:
                                attempt_second_security_type = False
                                order.set_status('REJECTED', account.log.info,
                                                 f'[{security_type}] Order would exceed maximum total strategy '
                                                 f'exposure for {qa_strategy_id}. Secondary security type will not be '
                                                 f'tried.')
                            if order.status != 'REJECTED' and not valid_daily_exposure:
                                attempt_second_security_type = False
                                order.set_status('REJECTED', account.log.info,
                                                 f'[{security_type}] Order would exceed maximum new daily exposure '
                                                 f'for {qa_strategy_id}. Secondary security type will not be '
                                                 f'tried.')

                        t, entry_order, child_exit_order = None, None, None
                        if order.status != 'REJECTED':
                            # 6. Get trade ID and other parameters and create Trade Record object
                            trade_id = trade_register.next_trade_id(broker)
                            order.order_ref.trade_id = trade_id
                            contract.exchange = 'SMART'  # avoids higher trade fees, see IB error 10311
                            strategy_model = deepcopy(strategy)
                            strategy_model.qa_model = order.qa_model_id  # strategy from config file has no model id
                            t = TradeRecord(ticker=order.ticker,
                                            trade_id=trade_id,
                                            strategy=strategy_model,
                                            market=US_MARKET,
                                            direction=strategy_model.direction,
                                            ib_contract_object=contract,
                                            entry_order_size=pos_size,
                                            entry_order_price=_order_price,
                                            account_id=account.id,
                                            account_alias=account.alias,
                                            norgate_asset_id=None,
                                            scheduled_entry_order_date=order.date)
                            t.setup_rank = order.order_rank
                            t.exposure = new_exposure_from_this_order

                            # 7. Create the order object (to be sent to IB)
                            if order.order_type_special and order.order_type_special:
                                if order.order_type_special not in TGL.supported_special_orders:
                                    order.set_status('REJECTED', account.log.error,
                                                     f'[{security_type}] Invalid OrderTypeSpecial: '
                                                     f'{order.order_type_special}. Valid '
                                                     f'special order types are '
                                                     f'{",".join([x for x in TGL.supported_special_orders])}')

                                elif order.order_type_special == 'Gap-Conditional':
                                    entry_order, child_exit_order = o.gap_and_go_special_order(contract,
                                                                                               order,
                                                                                               pos_size,
                                                                                               _order_price)

                                elif order.order_type_special == 'Child-MOC':
                                    entry_order = o.limit_order(action=order.action, quantity=pos_size,
                                                                limit_price=_order_price, order_ref=order.order_ref.ref,
                                                                ORTH=False, tif=order.tif)
                                    child_exit_order = o.moc_exit_order(parent_raw_entry_order=order, pos_size=pos_size)

                            else:
                                if order.order_type == 'LMT':
                                    entry_order = o.limit_order(action=order.action, quantity=pos_size,
                                                                limit_price=_order_price, order_ref=order.order_ref.ref,
                                                                ORTH=False, tif=order.tif)

                                elif order.order_type == 'MKT':
                                    entry_order = o.market_order(action=order.action, quantity=pos_size,
                                                                 order_ref=order.order_ref.ref, tif=order.tif)

                                else:
                                    # this code should be unreachable - order types errors are handled above.
                                    self.raise_oms_error((1, ValueError('Unhandled order type'), True))

                        if order.status not in ['INVALID', 'REJECTED']:

                            order.associated_trade = t

                            # 8. Add GoodTilTime if specified
                            if order.gtd_time:
                                entry_order.goodTillDate = order.gtd_time

                            # 10. Add IB ALGO to order, if applicable
                            if order.ib_algo == 'DARKICE':
                                display_size = int(entry_order.totalQuantity * self.config['DarkIce']/100)
                                entry_order = o.add_algo_dark_ice(entry_order, displaySize=display_size)

                            # 11. Save to trade register
                            entry_order.account = account.id
                            t.entry_order_record = o.OrderRecord(entry_order)
                            if child_exit_order:
                                child_exit_order.account = account.id
                                t.exit_order_records = [o.OrderRecord(child_exit_order)]
                            trade_register.append_trade(t)

                            # 12. Place Order
                            if child_exit_order:
                                account.log.info(f'Order {n}: Placing entry order for {pos_size} x {order.ticker} '
                                                 f'{security_type} via {t.contract.exchange}, with attached child '
                                                 f'exit order type: {child_exit_order.orderType}')
                                t.place_bracket_orders(broker)
                                account.log.info(f'Order {n}: Entry order placed with order Id '
                                                 f'{t.entry_order_record.order.orderId}')
                                account.log.info(f'Order {n}: Exit order placed with order Id '
                                                 f'{t.exit_order_records[0].order.orderId}')
                            else:
                                account.log.info(f'Order {n}: Placing entry order for {pos_size} x {order.ticker} '
                                                 f'{security_type} via {t.contract.exchange}')
                                t.place_entry_order(broker)
                                account.log.info(f'Order {n}: entry order placed with order Id '
                                                 f'{t.entry_order_record.order.orderId}')

                            # 13. Wait for broker response
                            # Background: All broker callbacks (whether order callbacks, or errors) are passed
                            # immediately to the relevant trade object (as identified by orderRef or orderId or permId).
                            # The trade object automatically processes these callbacks and updates its status
                            # accordingly.  Hence we watch the trade object here, not the broker.  To see how callbacks
                            # and errors are handled by the trade, see trades.TradeRecord.calculate_status() and
                            # trades.TradeRecord.process_ib_error()
                            if t.event_status_change.wait(30):

                                # the order has been placed and trade status has changed
                                t.event_status_change.clear()
                                sleeper.sleep(0.5)  # allow time for secondary callbacks & errors to be received
                                msg = '\n'.join(
                                    [f'Order {err.id}, Error: {err.code}: {err.string}' for err in t.errors
                                     if err.code not in self.config['Ignore IB Errors']])

                                if t.status in ['PLACED', 'OPEN']:
                                    if 404 in t.error_codes:
                                        # order placed, but shares not found for shorting
                                        broker.cancel_order(t.entry_order_record.order.orderId)
                                        t.set_status('CANCELLED')
                                        t.event_status_change.clear()
                                        t.cancel_reason.append(['Shares not found for shorting'])
                                        order.set_status(f'REJECTED', account.log.warning,
                                                         f'[{security_type}]  Shares not available for shorting')
                                    else:
                                        # the entry order has been successfully placed
                                        price = 'Market Price' if order.order_type == 'MKT' else f'${_order_price:.2f}'
                                        order.set_status('PLACED', account.log.info,
                                                         f'Order {n}: {order.action} {pos_size} x {order.ticker} '
                                                         f'{security_type} @ {price} successfully '
                                                         f'placed.')
                                        strategy.todays_exposure += new_exposure_from_this_order
                                        order.final_instrument_type = security_type
                                        break
                                elif t.status == 'CANCELLED':
                                    # order has been cancelled/rejected by IB.  The trade will be cancelled.
                                    order.set_status('REJECTED', account.log.warning,
                                                     f'[{security_type}] Order cancelled/rejected by broker. '
                                                     f'Broker error message: {msg}')
                            else:
                                # no callback was received from the broker. Cancel the trade.
                                broker.cancel_order(t.entry_order_record.order.orderId)
                                t.status = 'CANCELLED'
                                order.set_status('REJECTED', account.log.warning,
                                                 f'[{security_type}] Order rejected by broker - No callback received '
                                                 f'from broker after placing order.')
                            order.ib_errors = t.errors

                    # this code will only be reached if the order could not be placed
                    account.log.info(f'Order {n}: {security_type} type order could not be placed:')

                if order.status not in ['PLACED', 'OPEN', 'COMPLETE']:
                    account.log.info(f'Order {n}: Order could not be placed . Order processing complete.')

        return todays_entry_orders

    def calc_target_time(self, last_issued, target_time_of_day):

        target_datetime = None
        n = 0
        while not target_datetime:
            target_date = [s for s in US_MARKET.session_list if s.trading_date > last_issued]
            if target_date:
                target_date.sort(key=lambda s: s.trading_date)
                target_date = target_date[0]
                target_datetime = datetime.combine(target_date.trading_date, target_time_of_day).replace(
                    tzinfo=US_MARKET.zone_info)
            else:
                if n > 53:  # ~8 hours
                    self.raise_oms_error((2, ValueError('Could not determine next trade date. Click '
                                                        'Restart OMS button to try again'), True))
                elif n == 5:
                    # send email notification after 15 minutes
                    EMAIL_MANAGER.send_email(subject='OMS Connection Alert',
                                             body='The OMS has not been able to make a connection to TWS '
                                                  'for 15 minutes.  The OMS will continue to try for up '
                                                  'to 8 hours before raising an error',
                                             text_type='plain')
                elif n > 0:
                    LOG.warning('Could not get trade date data from broker.  Trying again in '
                                f'{min(n,  10)} minute{"s" if n > 1 else ""}.')
                    sleeper.sleep(min(n, 10) * 60)

                n += 1
                US_MARKET.update_market_sessions(self.data_broker)

        return target_datetime

    def issue_daily_reports(self, for_date=None, history_lookback_days=30):

        # Calling this function generates a Daily Report file which is emailed to the user for EACH ACCOUNT
        # The report contains
        # (1) summary performance data, per strategy
        # (2) new orders placed today
        # (3) open positions at the broker

        # The "Daily Report Time" in the config file should be set to a time later than "Trade Start Time" to ensure
        # that the daily report incorporates the most recent information.

        # Much of the code here has to do with formatting the output file.

        try:

            self._reporting_lock.acquire(blocking=True)
            self._blocking_function = 'daily_reports'

            if not for_date:
                for_date = US_MARKET.current_time.date()

            LOG.info('Generating daily reports')

            for account in self.trading_accounts:

                if account.ib_type == 'FUND ADVISORY' and not self.config['Report on Advisor Account']:
                    continue

                broker, trade_register = self.retrieve_broker(account, refresh_data=True)
                trades = [t for t in trade_register.trades if t.status != 'DRAFT']

                # DAILY SUMMARY REPORT

                file_name = f'{account.alias} - DAILY REPORT {for_date.strftime("%Y-%m-%d")}.xlsx'
                file_path = os.path.join(ADDR.folder_daily_reports, file_name)

                if trades:

                    LOG.info(f'Generating daily report for account {account.alias}')


                    # 1. Get current prices
                    n = 0
                    while n < 10 and [t for t in trades if not t.current_price if t.current_position != 0]:
                        n += 1
                        # get current stock prices
                        broker.update_current_prices(account.id)
                        sleeper.sleep(5)
                    missing_prices = [t for t in trades if not t.current_price if t.current_position != 0]
                    for t in missing_prices:
                        LOG.debug(f'Could not obtain current price for trade {t.trade_id}, Position of '
                                  f'{t.current_position} x {t.contract.symbol}, con Id {t.contract.conId}.')

                    # 2. Prepare summary strategy performance data
                    strategy_summary, new_trades, rejected_trades = [], [], []
                    strategy_numbers = list(set([t.strategy.qa_id for t in trades] +
                                                [k for k,v  in account.strategy_allocations.items() if v > 0]))
                    strategy_numbers.sort()
                    total_realised, total_unrealised, total_total = 0, 0, 0
                    for s in strategy_numbers:
                        all_prices_current = all(
                            [(t.current_price is not None
                              or t.current_position == 0
                              or not t.active) for t in trades if t.strategy.qa_id == s])
                        realised_pnl = sum([t.net_pnl for t in trades if t.strategy.qa_id == s])
                        total_realised += realised_pnl
                        if all_prices_current:
                            unrealised_pnl = (sum([t.unrealised_pnl for t in trades if t.strategy.qa_id == s])
                                              if all_prices_current else 'Incomplete Price Data')
                            total_pnl = realised_pnl + unrealised_pnl
                            total_unrealised += unrealised_pnl
                            total_total += total_pnl
                        else:
                            unrealised_pnl = 'Missing Current Prices'
                            total_pnl = ''
                        strategy_summary.append([s, realised_pnl, unrealised_pnl, total_pnl])
                    strategy_summary = [[f'Performance as at {datetime.strftime(for_date, "%Y-%m-%d")}', '', '', ''],
                                        ['Strategy', 'Realised PnL', 'Unrealised PnL', 'Total PnL']] + strategy_summary
                    strategy_summary.extend([['', '', '', ''],
                                             ['Total PnL', total_realised, total_unrealised, total_total]])

                    # 3. Write summary data to Strategy Performance sheet
                    wb = Workbook()
                    ws = wb.create_sheet(title='Strategy Performance')
                    for i, trade_data in enumerate(strategy_summary):
                        for k, value in enumerate(trade_data):
                            ws.cell(row=i + 1, column=k + 1).value = value
                            if k in [1, 2, 3]:
                                ws.cell(row=i + 1, column=k + 1).number_format = '$* #,##0;-$* #,##0'
                    ws.column_dimensions['A'].width = 20
                    ws.column_dimensions['B'].width = 14
                    ws.column_dimensions['C'].width = 14
                    ws.column_dimensions['D'].width = 14

                    # 4. Add Prior Day's Performance to Strategy Performance sheet
                    file_path_pattern = r'DAILY REPORT \d{4}-\d{2}-\d{2}\.xlsx$'
                    prior_report_files = [f for f in os.listdir(ADDR.folder_daily_reports)
                                          if ' - ' in f
                                          and f.split(' - ')[0] == account.alias
                                          and re.match(file_path_pattern, f.split(' - ')[1])
                                          and f[-15:] != file_name[-15:]]
                    sheet_insertion = []
                    if prior_report_files:
                        prior_report_file_name = max(prior_report_files)
                        prior_date = datetime.strptime(prior_report_file_name[-15:].replace('.xlsx', ''),
                                                       "%Y-%m-%d").date()
                        prior_wb_path = os.path.join(ADDR.folder_daily_reports, prior_report_file_name)
                        prior_wb = openpyxl.load_workbook(prior_wb_path)
                        prior_ws = prior_wb['Strategy Performance']

                        prior_strategy_summary = []
                        for row in prior_ws.iter_rows(min_row=3, max_row=prior_ws.max_row, max_col=4, values_only=True):
                            if row[0] in ['', 'Total PnL', None]:
                                break
                            prior_strategy_summary.append(list(row))

                        # calculate totals
                        prior_realised = sum([float(ps[1]) for ps in prior_strategy_summary if ps[1]])
                        prior_unrealised = sum([float(ps[2]) for ps in prior_strategy_summary if ps[2]])
                        prior_total = sum([float(ps[3]) for ps in prior_strategy_summary if ps[3]])


                        # add row per strategy
                        sheet_insertion = [[f'Performance as at {datetime.strftime(prior_date, "%Y-%m-%d")}',
                                           '', '', '', '', 'Change', 'Change'],
                                           ['Strategy', 'Realised PnL', 'Unrealised PnL', 'Total PnL', '',
                                            '$', '%']]
                        for today_strategy in strategy_summary[2:-2]:
                            prior_strategy = [s for s in prior_strategy_summary if s[0] == today_strategy[0]]
                            if prior_strategy:
                                prior_strategy = prior_strategy_summary.pop(
                                    prior_strategy_summary.index(prior_strategy[0]))
                                today_pnl, prior_pnl = float(today_strategy[3]), float(prior_strategy[3])
                                change_dollars = today_pnl - prior_pnl
                                change_pct = change_dollars/abs(prior_pnl) if prior_pnl else 'No data'
                                sheet_insertion.append(prior_strategy + ['', change_dollars, change_pct])
                            else:
                                sheet_insertion.append([today_strategy[0], 'No data', 'No data', 'No data', 'No data',
                                                        '', 'No data'])

                        for prior_strategy in prior_strategy_summary:  # handle any new strategies for this day
                            sheet_insertion.append(prior_strategy + ['', '-', '-'])

                        # add summary totals
                        total_change_dollar = sum([float(summary[5]) for summary in sheet_insertion[2:]
                                                   if summary[5] not in ['', '-', 'No data']])
                        total_change_pct = float(total_change_dollar) / abs(float(prior_total)) \
                            if prior_total else 'No data'
                        sheet_insertion.append(['', '', '', '', '', '', ''])
                        sheet_insertion.append(['Total PnL', prior_realised, prior_unrealised, prior_total,
                                                '', total_change_dollar, total_change_pct])

                        for i, prior_ss in enumerate(sheet_insertion):
                            for k, value in enumerate(prior_ss):
                                ws.cell(row=i + 1, column=k + 6).value = value
                                if k in [1, 2, 3, 5]:
                                    ws.cell(row=i + 1, column=k + 6).number_format = '$* #,##0;-$* #,##0'
                                elif k in [6]:
                                    ws.cell(row=i + 1, column=k + 6).number_format = '0.00%'
                                    ws.cell(row=i + 1, column=k + 6).alignment = Alignment(horizontal='center')

                        ws.column_dimensions['F'].width = 20
                        for col in ['G', 'H', 'I', 'K', 'L']:
                            ws.column_dimensions[col].width = 14

                    else:
                        ws.cell(row=1, column=6).value = 'No data for prior date available.'

                    for cell in ws[1]:  # Header
                        cell.font = Font(bold=True)
                    for cell in ws[2]:  # Column headers
                        cell.font = Font(bold=True)

                    # 5. Write NAVs to Strategy Performance page
                    start_row = max([len(strategy_summary), len(sheet_insertion)]) + 4
                    nav_summary = [['Account', 'NAV', 'Since Prior', 'This Month', 'This Year', 'Last Updated']]
                    for account_id, nav_detail in self._nav_monitor.current_navs.items():
                        nav_summary.append([account_id] + nav_detail)

                    for i, nav in enumerate(nav_summary):
                        for k, value in enumerate(nav):
                            ws.cell(row=start_row + i, column=k + 1).value = value
                            if k == 1:
                                ws.cell(row=start_row + i, column=k + 1).number_format = '[$AUD] #,##0;-[$AUD] #,##0'
                            elif k in [2, 3, 4]:
                                ws.cell(row=start_row + i, column=k + 1).number_format = '0.00%'

                    # 5. Write all open orders to Active Orders sheet
                    ws = wb.create_sheet(title='Active Orders')

                    o_data = [['OrderRef', 'Symbol', 'Sec Type', 'Order Size', 'Order Price', 'Order Value',
                               'Setup Score', 'Order ID', 'IB Error Codes']]
                    all_errors = []
                    if account.id in broker.api.raw_open_orders.keys():
                        for orderRef, order_data in broker.api.raw_open_orders[account.id].items():
                            order_ref = o.decodeOrderRef(orderRef)
                            t = broker.api.find_trade_by_orderRef(order=order_data[2])[0]
                            if order_ref.trade_id != 0 and t:  # filter out manual (non OMS) trades
                                # Raw Open Order data structure:
                                order_id, contract, order, order_state = order_data[0], order_data[1], order_data[2], \
                                                                         order_data[3]
                                errors = [e for e in t.errors
                                          if e.id == order_id and
                                          e.code not in self.config['Ignore IB Errors']]
                                all_errors.extend(errors)
                                o_data = o_data + [[order_ref.ref, contract.symbol, contract.secType,
                                                   order.totalQuantity, order.lmtPrice,
                                                   order.totalQuantity * Decimal(str(order.lmtPrice)),
                                                   t.setup_rank, order_id,
                                                   ','.join([str(e.code) for e in errors])]]

                    for i, trade_data in enumerate(o_data):
                        for k, value in enumerate(trade_data):
                            ws.cell(row=i + 1, column=k + 1).value = value
                            if k == 4:
                                ws.cell(row=i + 1, column=k + 1).number_format = '$* #,##0.00;-$* #,##0.00'
                            if k == 5:
                                ws.cell(row=i + 1, column=k + 1).number_format = '$* #,##0;-$* #,##0'
                    ws.column_dimensions['A'].width = 20
                    for c in [chr(i) for i in range(ord('B'), ord('H') + 1)]:
                        ws.column_dimensions[c].width = 12
                    for cell in ws[1]:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='top')

                    # 6. Write IB Errors for today's entries to IB Error Codes sheet
                    ws = wb.create_sheet(title='IB Error Codes')
                    all_errors = [['Order', 'Code', 'Message'] + [e.id, e.code, e.string] for e in all_errors]

                    for i, v in enumerate(all_errors):
                        ws.cell(row=i + 1, column=1).value = v[0]
                        ws.cell(row=i + 1, column=2).value = v[1]
                        ws.cell(row=i + 1, column=3).value = v[2]

                    # 7. Write trade register to Trade Register sheet
                    real_trades = [t for t in trades
                                   if t.status in ['OPEN', 'COMPLETE']
                                   and (not history_lookback_days or
                                        t.created_datetime.date() >= (datetime.now() -
                                                                      timedelta(days=history_lookback_days)).date())]
                    real_trades.sort(key=lambda x: x.created_datetime, reverse=True)
                    self.add_trade_history_sheet(wb, real_trades)

                    # 8. Note any partial exits
                    partial_exits = [t for t in trades if t.current_position not in [0, t.achieved_entry_size]]
                    partial_exits_text = ''
                    if partial_exits:
                        partial_exits_text = '\n\nPARTIAL EXIT RECORDED FOR: ' + ', '.join(
                            [t.ticker for t in partial_exits])

                    # 9. Save and email report
                    del wb['Sheet']
                    self.save_workbook(wb, file_path)
                    wb.close()
                    email_body = f"Attached is the daily report for {account.alias} for market date {for_date}" \
                                 + partial_exits_text

                else:
                    file_path = None
                    email_body = f"No trade activity or current positions to report for {account.alias} for " \
                                 f"market date {for_date}"
                    LOG.info(email_body)

                # email report, if applicable
                if email_body:
                    EMAIL_MANAGER.send_email(subject=f'{account.alias} - Daily Report '
                                                     f'{datetime.strftime(for_date, "%Y%m%d")}',
                                             body=email_body,
                                             text_type='plain',
                                             attachments_file_paths=([file_path]
                                                                     if file_path and file_path is not None
                                                                     else [])
                                             )

                    account.log.info('Daily Report Issued\n')

        except Exception as e:
            # self.raise_oms_error((1, e, True))
            raise e
        finally:
            self._blocking_function = ''
            self._reporting_lock.release()

    def export_trade_history(self):

        # the user calls this function to generate a simple trade history file

        self._reporting_lock.acquire(blocking=True)
        self._blocking_function = 'trade_history'

        LOG.info('Exporting trade history to Excel')
        # Deprecated - Excel setup
        # wb = Workbook()
        # file_name = f'TRADE HISTORY TO {datetime.strftime(datetime.now(), "%Y-%m-%d")}.xlsx'
        # file_path = os.path.join(ADDR.folder_output_trade_exports, file_name)

        for account in self.trading_accounts:

            if account.ib_type == 'FUND ADVISORY' and not self.config['Report on Advisor Account']:
                continue

            broker, trade_register = self.retrieve_broker(account, refresh_data=True)
            trades = trade_register.trades

            LOG.info(f'Generating trade history for account {account.alias}')
            if [t for t in trades if not t.current_price]:
                # get current stock prices
                broker.update_current_prices(account.id)
                sleeper.sleep(5)

            # write trade register
            real_trades = [t for t in trades
                           if t.status in ['OPEN', 'COMPLETE']]
            real_trades.sort(key=lambda x: x.created_datetime, reverse=True)

            # Current - export all accounts to separate CSV files
            file_name = f'TRADE_HISTORY_{account.alias}_{datetime.strftime(datetime.now(), "%Y-%m-%d")}.xlsx'
            file_path = os.path.join(ADDR.folder_output_trade_exports, file_name)
            wb = Workbook()
            self.add_trade_history_sheet(wb, real_trades, title='Trade History')
            for sheet_name in wb.sheetnames:
                if sheet_name in ['Sheet', 'Sheet1']:
                    wb.remove(wb[sheet_name])
            try:
                wb.save(file_path)
                wb.close()
            except PermissionError:
                LOG.warning('Could not save Trade History file - check file is not already open.')

        self._blocking_function = ''
        self._reporting_lock.release()

        # SIGNALS.trade_history_generated.emit(file_path)  # deprecated
        SIGNALS.trade_history_generated.emit(ADDR.folder_output_trade_exports)

    @staticmethod
    def add_trade_history_sheet(workbook, trades, title='Trade Register'):
        ws = workbook.create_sheet(title=title)
        t_data = [['Strategy', 'Symbol', 'Side', 'DateIn', 'TimeIn', 'QtyIn', 'PriceIn',
                   'DateOut', 'TimeOut', 'QtyOut', 'PriceOut', 'FeesIn', 'FeesOut', 'Status',
                   'Order Price', 'Order Size', 'Entry Value', 'Current Price', 'Exit Type', 'Realised PnL',
                   'Unrealised PnL']] + \
                 [[t.strategy.qa_strategy_model, t.ticker, t.direction.upper(),
                   t.entry_datetime.replace(tzinfo=None).date() if t.entry_datetime else '-',
                   t.entry_datetime.replace(tzinfo=None).time() if t.entry_datetime else '-',
                   abs(t.achieved_entry_size), t.avg_entry_price,
                   t.exit_datetime.replace(tzinfo=None).date() if t.exit_datetime else '-',
                   t.exit_datetime.replace(tzinfo=None).time() if t.exit_datetime else '-',
                   abs(t.achieved_exit_size), t.avg_exit_price, t.brokerage_entry, t.brokerage_exit,
                   t.status, t.order_price, t.order_size, t.entry_value, t.current_price, t.exit_type,
                   t.net_pnl if t.status != 'OPEN' else '',
                   t.unrealised_pnl if t.status == 'OPEN' else '']
                  for t in trades]
        for i, trade_data in enumerate(t_data):
            for k, value in enumerate(trade_data):
                ws.cell(row=i + 1, column=k + 1).value = value
                ws.cell(row=i + 1, column=k + 1).alignment = Alignment(vertical='top')
                if k in [1, 2]:
                    ws.cell(row=i + 1, column=k + 1).alignment = Alignment(horizontal='center')
                elif k in [6, 10, 11, 12, 14, 17, 19, 20]:
                    ws.cell(row=i + 1, column=k + 1).number_format = '$* #,##0.00;-$* #,##0.00'
                elif k in [16]:
                    ws.cell(row=i + 1, column=k + 1).number_format = '$* #,##0;-$* #,##0'
                elif k in [3, 7]:
                    if value:
                        ws.cell(row=i + 1, column=k + 1).number_format = 'd/m/yyyy'
                elif k in [4, 8]:
                    if value:
                        ws.cell(row=i + 1, column=k + 1).number_format = 'h:mm:ss AM/PM'
        ws.column_dimensions['A'].width = 20
        for c in [chr(i) for i in range(ord('B'), ord('N') + 1)]:
            ws.column_dimensions[c].width = 12
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, horizontal='center')

    def issue_rt_trade_list(self, for_date=None):

        # Calling this function generates an RT Trade List file which is saved for EACH ACCOUNT
        # The file is intended as an import into RealTest, to enable generation of the next day's orders

        # The "RT Output File Time" in the config file should be set to a time earlier than "Trade Start Time" to
        # ensure sufficient time for this output file to be used by RealTest to generate the new Orders File

        # Only one report is generated per day.  To generate the report again, delete the existing report.

        if not for_date:
            for_date = US_MARKET.current_time.date()

        self._reporting_lock.acquire(blocking=True)
        self._blocking_function = 'rt_output'

        for account in self.trading_accounts:

            if account.ib_type == 'FUND ADVISORY' and not self.config['Report on Advisor Account']:
                continue

            broker, trade_register = self.retrieve_broker(account, refresh_data=True)
            trades = trade_register.trades

            csv_file_name = f'{account.alias} trade list.csv'
            csv_file_path = os.path.join(ADDR.folder_output_trade_lists, csv_file_name)

            archive_csv_file_name = f'{account.alias} trade list {datetime.strftime(for_date, "%Y-%m-%d")}.csv'
            archive_csv_file_path = os.path.join(ADDR.folder_output_trade_lists, 'Archive', archive_csv_file_name)

            LOG.info(f'Generating RT trade list for {account.alias} for {datetime.strftime(for_date, "%Y%m%d")}')

            self.reconcile(broker, account)

            wb = Workbook()

            # write all open positions
            ws = wb.create_sheet(title=f'{account.alias} Trade List')
            open_positions = [t for t in trades if t.current_position != 0]
            open_positions.sort(key=lambda x: x.created_datetime, reverse=True)
            # Background/logic here for handling partial exits:
            # - RT does not handle partial exits within one 'line'.  It requires us to split a partially exited
            # trade into 2 trades - the exited component (as a closed trade) and remaining position (as an open trade)
            # - Additionally, for our purposes, we do not require any exit data to be reported in the RT output file,
            # so (1) we don't report any exit data (blanks in list below), and (2) we don't report the exited
            # component of a partially exited trade - we just report the open position as a stand-alone trade.
            # cf https://forum.mhptrading.com/t/handling-missed-exits-in-orderclerk-strategies/1514/12
            # RT & POSITION SIZE:
            # RT does not handle negative positions (it just uses the Side).  Hence report abs(t.current_position)
            t_data = [['Symbol', 'Strategy', 'Side', 'Shares', 'DateIn', 'TimeIn',
                       'QtyIn', 'PriceIn', 'FeesIn', 'DateOut', 'TimeOut', 'QtyOut', 'PriceOut', 'FeesOut']] + \
                     [[t.ticker, t.strategy.qa_strategy_model, t.direction, abs(t.current_position),
                       t.entry_datetime.date() if t.entry_datetime else None,
                       t.entry_datetime.time() if t.entry_datetime else None,
                       abs(t.current_position), t.avg_entry_price,
                       t.brokerage_entry,
                       '', '', '', '', '']
                      for t in open_positions]
            for i, trade_data in enumerate(t_data):
                for k, value in enumerate(trade_data):
                    ws.cell(row=i + 1, column=k + 1).value = value
            for c in [chr(i) for i in range(ord('A'), ord('K') + 1)]:
                ws.column_dimensions[c].width = 12

            del wb['Sheet']

            for file_path in [csv_file_path, archive_csv_file_path]:
                with open(file_path, 'w', newline='') as csvfile:
                    csvwriter = csv.writer(csvfile)
                    for row in ws.iter_rows():
                        csvwriter.writerow([cell.value for cell in row])

            wb.close()

            if trades:
                account.log.info(f'{account.alias} - Trade List (for use in RT) saved at {csv_file_path}\n')
            else:
                account.log.info(f"{account.alias} - No trades found - empty RT trade list saved at {csv_file_path}\n")

        self._blocking_function = ''
        self._reporting_lock.release()

    def save_workbook(self, wb, filepath):
        success, n, error_raised = False, 0, False
        while not success:
            try:
                wb.save(filepath)
                success = True
            except PermissionError:
                if not error_raised:
                    self.raise_oms_error((3, PermissionError(f'Could not access the file {filepath}.  '
                                                             f'Is it open?'), False))
                error_raised = True
                sleeper.sleep(5)

    @DeprecationWarning  # migrated to email_manager module
    def email_notification(self, email_subject, email_body, attachment_file_path=None):

        sender_address = self.config['Email Sending Address']
        recipient_address = self.config['Email Recipients']
        username = self.config['Email Username']
        password = self.config['Email Password']
        host = self.config['Email Host Address']
        smtp_port = self.config['SMTP Port']

        msg = MIMEMultipart()
        msg['From'] = formataddr((str(Header('QA OMS Info', 'utf-8')), sender_address))
        msg['To'] = recipient_address
        msg['Subject'] = email_subject

        msg.attach(MIMEText(email_body, 'html'))

        if attachment_file_path is not None and os.path.isfile(attachment_file_path):
            with open(attachment_file_path, "r") as fil:
                ext = attachment_file_path.split('.')[-1]
                if ext in ['xls', 'xlsx']:
                    attachment = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    attachment.set_payload(open(attachment_file_path, 'rb').read())
                    encoders.encode_base64(attachment)
                    attachment.add_header('Content-Disposition', 'attachment',
                                          filename=os.path.basename(attachment_file_path))
                    msg.attach(attachment)
                else:
                    attached_file = MIMEApplication(fil.read(), _subtype=ext)
                    attached_file.add_header(
                        'content-disposition', 'attachment', filename=basename(attachment_file_path))
                    msg.attach(attached_file)

        success, n = False, 0
        while not success:
            self.stop_signal()
            n += 1
            try:
                server = SMTP(host, smtp_port)
                server.starttls()
                server.login(username, password)
                server.sendmail(sender_address, recipient_address, msg.as_string())
                server.quit()
                success = True
            except (SMTPDataError, SMTPServerDisconnected) as exc:
                if n < 30:
                    wait_time = min(n, 12) ** 3
                    self.raise_oms_error((3,
                                                 ConnectionError(f'Email attempt {n} failed: {exc}.\n'
                                                                 f'Waiting {util.timer_display_text(wait_time)} '
                                                                 f'to try again'), False))
                    sleeper.sleep(wait_time)
                else:
                    self.raise_oms_error((2,
                                                  ConnectionError(f'Could not send email after {n} attempts. Address '
                                                                  f'the issue and then click the Restart OMS button '
                                                                  f'to try again'), True))
            except SMTPAuthenticationError:
                self.raise_oms_error((2, ConnectionError('Could not authenticate email connection - please '
                                                         'check config settings and restart the OMS'), True))

    def cancel_open_orders(self):

        self._reporting_lock.acquire(blocking=True)

        # Iterate over each account to cancel entries and exits
        for accnt in self.trading_accounts:

            self.stop_signal()

            self.order_processing_status = f'Cancelling orders for {accnt.alias}'
            LOG.warning(f"\n\nCancelling orders for {accnt.alias}\n")

            # get broker and trader register objects for this account
            LOG.info('Refreshing trade data from broker.')
            broker, trade_register = self.retrieve_broker(accnt, refresh_data=True)

            if accnt.id in broker.api.raw_open_orders.keys() and broker.api.raw_open_orders[accnt.id]:
                self.stop_signal()
                trades_to_delete = []
                raw_open_orders = deepcopy(list(broker.api.raw_open_orders[accnt.id].items()))  # may change size during iteration
                for orderRef, order_data in raw_open_orders:
                    order_id, contract, order, order_state = order_data[0], order_data[1], order_data[2], \
                        order_data[3]
                    order_ref = o.decodeOrderRef(orderRef)
                    t = broker.api.find_trade_by_orderRef(order)[0]
                    if order_ref.trade_id != 0 and t:
                        if 'ENTRY' in order_ref.ref:
                            if t.status.upper() not in ['OPEN', 'COMPLETE'] and t.current_position == 0:
                                broker.cancel_order(order_id)
                                accnt.log.info(f'Cancelled entry order {order_id} ({order_ref.ref}). '
                                               f'Trade {t.trade_id} deleted.')
                                trades_to_delete.append(t)
                            else:
                                accnt.log.warning(f'Could not cancel entry order {order_id} ({order_ref.ref}) '
                                                  f'for trade {t.trade_id} - position already open.')
                        else:
                            broker.cancel_order(order_id)
                            eor = [e for e in t.exit_order_records if e.order.orderRef == order_ref.ref][0]
                            t.exit_order_records.remove(eor)
                            accnt.log.info(f'Cancelled exit order {order_id} ({order_ref.ref}).')
                for t in trades_to_delete:
                    trade_register.trades.remove(t)
            else:
                print(f'No orders were found to be cancelled')

        self.order_processing_status = ''
        self._reporting_lock.release()

    @staticmethod
    def plo(order_list):
        # little snippet of code to return the correctly pluralised version of order/orders
        if isinstance(order_list, list):
            count = len(order_list)
            if count == 1:
                return 'order'
            else:
                return 'orders'
        else:
            return ''

    def shutdown(self):

        self.keep_alive = False

        LOG.info('Disconnecting API')
        from src.ib_api.interface.ib_api import IB_API
        for api in IB_API._instances:
            # disconnect from broker
            api.keep_alive = False
            api.disconnect()

        # turn off autosave and save trade registers
        LOG.info('Saving trade register(s)')
        for tr in TradeRegister._instances.values():
            tr.enable_autosave = False

        # stop email handler thread
        LOG.info('Stopping emailer')
        for smtp_handler in [h for h in LOG.handlers if type(h) == customSMTPHandler]:
            # Deprecated
            smtp_handler.keep_alive = False
        EMAIL_MANAGER.keep_alive = False

        # stop NAV monitor
        self._nav_monitor.keep_alive = False

        self._stop_signal = False  # release the lock, so that the OMS can exit

        LOG.info('Stopping threads')
        kill_threads = True
        while kill_threads:
            kill_threads = [t for t in threading.enumerate() if
                            any([x in t.name for x in ['Trade Register Autosave', 'Emailing Thread', 'OMS Thread',
                                                       'NAV Monitor']])]
            for thread in kill_threads:
                thread.join(0.2)

        LOG.debug('All OMS threads joined')


def main():
    from src.gui.splash_screen import QASplashScreen
    
    app = QApplication(sys.argv)

    pixmap = QPixmap(500, 300)
    splash = QASplashScreen(
        pixmap=pixmap,
        software_name="Quantive Alpha\nOrder Management System",
        version_info=__version__
    )
    splash.show()

    oms = OMS()
    oms_gui = OMSgui(oms)
    oms_gui.show()
    splash.finish(oms_gui)
    app.exec()


if __name__ == '__main__':
    main()
